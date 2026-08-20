from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path

import numpy as np
import pytest

from sensorarray_backend.core.session_data import (
    SESSION_SCHEMA_VERSION,
    SessionFrame,
    export_session_bytes,
    frames_to_measurement_ascii_bytes,
    load_session_frames,
    session_model_from_payload,
)


def _session_frame(
    mode: str,
    *,
    rows: int,
    seq: int,
    boot_id: int = 7,
    rail_boot_id: int | None = None,
) -> SessionFrame:
    modes = {
        "CAP": ("CAP",) * 8,
        "VOLT": ("VOLT",) * 8,
        "RES": ("RES",) * 8,
        "MIXED": tuple(("CAP", "VOLT", "RES")[index % 3] if index < rows else "NONE" for index in range(8)),
    }[mode]
    units = tuple({"CAP": "pF", "VOLT": "V", "RES": "ohm", "NONE": ""}[item] for item in modes)
    scales = tuple({"CAP": -6, "VOLT": -6, "RES": -3, "NONE": None}[item] for item in modes)
    physical = np.full(64, np.nan, dtype=np.float64)
    raw = np.full(64, np.nan, dtype=np.float64)
    valid = np.zeros(64, dtype=bool)
    fresh = np.zeros(64, dtype=bool)
    expected = np.zeros(64, dtype=bool)
    acquired = np.zeros(64, dtype=bool)
    error = np.zeros(64, dtype=bool)
    codes = np.full(64, -1, dtype=np.int16)
    reasons = np.full(64, "", dtype=object)
    pga = np.full(64, -1, dtype=np.int16)
    bypass = np.zeros(64, dtype=bool)
    for row in range(rows):
        row_mode = modes[row]
        for col in range(8):
            index = row * 8 + col
            expected[index] = True
            acquired[index] = True
            fresh[index] = True
            valid[index] = True
            if row_mode == "CAP":
                raw[index] = 39_000_000 + index * 1_000
                physical[index] = raw[index] * 1e-6 - 33.0
            elif row_mode == "VOLT":
                raw[index] = -1_000_000 + index * 10_000
                physical[index] = raw[index] * 1e-6
                pga[index] = 0 if col == 0 else 1
                bypass[index] = col == 0
            else:
                raw[index] = 10_000_000 + index * 1_000
                physical[index] = raw[index] * 1e-3
                pga[index] = 2

    # Acquired OPEN remains fresh but invalid. A non-acquired cell remains
    # explicitly expected, not fresh, invalid, and carries the firmware's
    # unsupported/not-acquired X14 representation.
    if mode in {"VOLT", "RES", "MIXED"}:
        open_index = next(
            (index for index in range(rows * 8) if modes[index // 8] in {"VOLT", "RES"}),
            None,
        )
        if open_index is not None:
            valid[open_index] = False
            error[open_index] = True
            codes[open_index] = 0x0D
            reasons[open_index] = "Open circuit"
            physical[open_index] = np.nan
            raw[open_index] = np.nan
            missing_index = open_index + 1
            acquired[missing_index] = False
            fresh[missing_index] = False
            valid[missing_index] = False
            error[missing_index] = True
            codes[missing_index] = 0x14
            reasons[missing_index] = "Unsupported measurement"
            physical[missing_index] = np.nan
            raw[missing_index] = np.nan

    configured = tuple(item if item != "NONE" else "CAP" for item in modes)
    wire = "".join({"CAP": "C", "VOLT": "V", "RES": "R", "NONE": "N"}[item] for item in modes)
    rail = {
        "railValid": True,
        "railFresh": True,
        "railAge": 2,
        "avddUv": 3_391_000,
        "avssUv": -2_500_000,
        "railSpanUv": 5_891_000,
        "railSource": "frame",
        "railReason": "ok",
        "bootId": boot_id if rail_boot_id is None else rail_boot_id,
    }
    generation = 3 + seq
    request_id = 40 + seq
    unit = {"CAP": "pF", "VOLT": "V", "RES": "ohm", "MIXED": ""}[mode]
    scale = {"CAP": -6, "VOLT": -6, "RES": -3, "MIXED": 0}[mode]
    return SessionFrame(
        seq=seq,
        timeSeconds=1000.0 + seq,
        rows=rows,
        valuesPf=physical if mode == "CAP" else np.full(64, np.nan),
        valid=valid,
        source="serial",
        measurementMode=mode,
        unit=unit,
        scale=scale,
        physicalValues=physical,
        rawFixed=raw,
        fresh=fresh,
        errorCodes=codes,
        pga=pga,
        generation=generation,
        requestId=request_id,
        sessionId="session-v3-test",
        connectionGeneration=2,
        bootId=boot_id,
        deviceTimestampUs=998_000 + seq,
        hostReceivedUtc=f"2026-08-20T00:00:{seq:02d}+00:00",
        hostWallTime=1_776_000_000.0 + seq,
        hostReceivedMonotonicNs=9_000_000 + seq,
        frameKind=mode,
        expected=expected,
        acquired=acquired,
        expectedKnown=np.ones(64, dtype=bool),
        acquiredKnown=np.ones(64, dtype=bool),
        freshKnown=np.ones(64, dtype=bool),
        error=error,
        errorReasons=reasons,
        pgaBypass=bypass,
        rowsGeneration=generation if mode in {"CAP", "MIXED"} else None,
        rowsRequestId=request_id if mode in {"CAP", "MIXED"} else None,
        modeGeneration=generation if mode in {"VOLT", "RES"} else None,
        modeRequestId=request_id if mode in {"VOLT", "RES"} else None,
        profileGeneration=generation if mode == "MIXED" else None,
        profileRequestId=request_id if mode == "MIXED" else None,
        configuredRowProfile=configured,
        wireRowProfile=wire if mode == "MIXED" else None,
        rowModes=modes,
        rowUnits=units,
        rowScales=scales,
        rail=rail,
    )


def _frame_payload(frame: SessionFrame) -> dict:
    def numbers(values: np.ndarray) -> list[float | None]:
        return [float(value) if np.isfinite(value) else None for value in np.asarray(values).reshape(64)]

    return {
        "schemaVersion": SESSION_SCHEMA_VERSION,
        "sessionId": frame.sessionId,
        "connectionGeneration": frame.connectionGeneration,
        "bootId": frame.bootId,
        "deviceTimestampUs": frame.deviceTimestampUs,
        "hostReceivedUtc": frame.hostReceivedUtc,
        "hostWallTime": frame.hostWallTime,
        "hostReceivedMonotonicNs": frame.hostReceivedMonotonicNs,
        "timeSeconds": frame.timeSeconds,
        "seq": frame.seq,
        "rows": frame.rows,
        "frameKind": frame.mode,
        "measurementMode": frame.mode,
        "unit": frame.unit,
        "scale": frame.scale,
        "generation": frame.generation,
        "requestId": frame.requestId,
        "rowsGeneration": frame.rowsGeneration,
        "rowsRequestId": frame.rowsRequestId,
        "modeGeneration": frame.modeGeneration,
        "modeRequestId": frame.modeRequestId,
        "profileGeneration": frame.profileGeneration,
        "profileRequestId": frame.profileRequestId,
        "configuredRowProfile": list(frame.configuredRowProfile or ()),
        "wireRowProfile": frame.wireRowProfile,
        "rowModes": list(frame.row_mode_values()),
        "rowUnits": list(frame.row_unit_values()),
        "rowScales": list(frame.row_scale_values()),
        "physicalValues": numbers(frame.physical_values()),
        "valuesPf": numbers(frame.valuesPf),
        "rawFixed": numbers(frame.raw_fixed_values()),
        "valid": frame.valid_matrix().reshape(64).tolist(),
        "fresh": frame.fresh_values().tolist(),
        "freshKnown": frame.fresh_known_values().tolist(),
        "expected": frame.expected_values().tolist(),
        "expectedKnown": frame.expected_known_values().tolist(),
        "acquired": frame.acquired_values().tolist(),
        "acquiredKnown": frame.acquired_known_values().tolist(),
        "error": frame.error_values().tolist(),
        "errorCodes": frame.error_code_values().tolist(),
        "errorReasons": frame.error_reason_values().tolist(),
        "pga": frame.pga_values().tolist(),
        "pgaBypass": frame.pga_bypass_values().tolist(),
        "rail": dict(frame.rail or {}),
        "source": frame.source,
    }


def _payload(frames: list[SessionFrame]) -> dict:
    latest = frames[-1]
    return {
        "metadata": {"schemaVersion": 3, "sessionId": "session-v3-test"},
        "display": {"voltageReference": "vss_relative"},
        "offsetsPf": [[0.0] * 8 for _ in range(8)],
        "currentMatrix": {
            "values": np.where(np.isfinite(latest.values_matrix()), latest.values_matrix(), np.nan).tolist(),
            "valid": latest.valid_matrix().tolist(),
        },
        "historyFrames": [_frame_payload(frame) for frame in frames],
        "rawLogs": [{"tag": "BOOT", "rawText": "BOOT,bootId=7"}],
        "events": [
            {"kind": "TRANSPORT_RECONNECT", "connectionGeneration": 2},
            {"kind": "DEVICE_REBOOT", "oldBootId": 6, "newBootId": 7},
        ],
    }


@pytest.mark.parametrize("fmt", ["csv", "xlsx", "mat", "h5", "zip"])
def test_v3_export_import_round_trip_preserves_domains_masks_errors_and_identity(tmp_path: Path, fmt: str):
    frames = [
        _session_frame("CAP", rows=8, seq=1),
        _session_frame("VOLT", rows=2, seq=2),
        _session_frame("RES", rows=1, seq=3),
        _session_frame("MIXED", rows=8, seq=4),
    ]
    data, _media_type, extension = export_session_bytes(_payload(frames), fmt)
    path = tmp_path / f"session.{extension}"
    path.write_bytes(data)
    loaded = load_session_frames(path)

    assert [frame.mode for frame in loaded] == ["CAP", "VOLT", "RES", "MIXED"]
    assert [frame.rows for frame in loaded] == [8, 2, 1, 8]
    for expected_frame, actual in zip(frames, loaded, strict=True):
        np.testing.assert_allclose(actual.physical_values(), expected_frame.physical_values(), equal_nan=True)
        assert actual.row_mode_values() == expected_frame.row_mode_values()
        assert actual.row_unit_values() == expected_frame.row_unit_values()
        assert actual.row_scale_values() == expected_frame.row_scale_values()
        np.testing.assert_array_equal(actual.expected_values(), expected_frame.expected_values())
        np.testing.assert_array_equal(actual.acquired_values(), expected_frame.acquired_values())
        np.testing.assert_array_equal(actual.fresh_values(), expected_frame.fresh_values())
        np.testing.assert_array_equal(actual.error_code_values(), expected_frame.error_code_values())
        assert actual.bootId == expected_frame.bootId
        assert actual.connectionGeneration == expected_frame.connectionGeneration
        assert actual.deviceTimestampUs == expected_frame.deviceTimestampUs
        assert actual.profileGeneration == expected_frame.profileGeneration
        assert (actual.rail or {}).get("avssUv") == -2_500_000

    voltage = loaded[1]
    open_index = 0
    assert voltage.acquired_values()[open_index]
    assert voltage.fresh_values()[open_index]
    assert not voltage.valid_matrix().reshape(64)[open_index]
    assert voltage.error_code_values()[open_index] == 0x0D
    assert voltage.expected_values()[1] and not voltage.acquired_values()[1] and not voltage.fresh_values()[1]


@pytest.mark.parametrize("rows", range(1, 9))
def test_mixed_v3_supports_every_rows_geometry_and_exact_n_suffix(rows: int):
    frame = _session_frame("MIXED", rows=rows, seq=rows)
    wire = frames_to_measurement_ascii_bytes([frame]).decode("ascii")
    expected_suffix = "N" * (8 - rows)
    assert f"profile={frame.wireRowProfile[:rows]}{expected_suffix}" in wire
    assert wire.count("\nMR,") == rows


def test_v2_missing_acquisition_and_freshness_migrate_to_unknown_not_valid_copy():
    values = [1.0] * 64
    payload = {
        "metadata": {"schemaVersion": 2},
        "currentMatrix": {"values": [[1.0] * 8 for _ in range(8)], "valid": [[True] * 8 for _ in range(8)]},
        "historyFrames": [
            {"seq": 1, "rows": 8, "measurementMode": "CAP", "valuesPf": values, "valid": [True] * 64}
        ],
    }
    frame = session_model_from_payload(payload).frames[0]
    assert frame.valid_matrix().all()
    assert not frame.expected_known_values().any()
    assert not frame.acquired_known_values().any()
    assert not frame.fresh_known_values().any()
    assert not frame.expected_values().any()
    assert not frame.acquired_values().any()


def test_zip_bundle_wide_tables_split_mixed_domains_and_keep_one_frame_per_row():
    frame = _session_frame("MIXED", rows=8, seq=4)
    data, media_type, extension = export_session_bytes(_payload([frame]), "zip")
    assert (media_type, extension) == ("application/zip", "zip")
    with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
        expected_names = {
            "metadata.csv", "frame_metadata.csv", "cap_wide.csv", "volt_wide.csv",
            "res_wide.csv", "quality_long.csv", "events.csv", "logs.csv", "cap_offsets.csv",
            "volt_vss_wide.csv", "volt_normalized_wide.csv",
        }
        assert expected_names.issubset(set(archive.namelist()))
        domain_rows = {}
        for name in ("cap_wide.csv", "volt_wide.csv", "res_wide.csv"):
            rows = list(csv.DictReader(io.StringIO(archive.read(name).decode("utf-8"))))
            assert len(rows) == 1
            domain_rows[name] = rows[0]
        assert domain_rows["cap_wide.csv"]["S1D1"] != ""
        assert domain_rows["volt_wide.csv"]["S1D1"] == ""
        assert domain_rows["res_wide.csv"]["S1D1"] == ""
        assert domain_rows["volt_wide.csv"]["S2D3"] != ""
        assert domain_rows["cap_wide.csv"]["S2D3"] == ""


def test_mat_h5_derived_voltage_requires_fresh_same_boot_rail(tmp_path: Path):
    same_boot = _session_frame("VOLT", rows=2, seq=1, boot_id=7, rail_boot_id=7)
    old_boot = _session_frame("VOLT", rows=2, seq=2, boot_id=8, rail_boot_id=7)
    payload = _payload([same_boot, old_boot])

    mat_data, _, _ = export_session_bytes(payload, "mat")
    from scipy.io import loadmat

    mat = loadmat(io.BytesIO(mat_data))
    assert np.isfinite(mat["volt_values_vss_relative"][0]).any()
    assert np.isnan(mat["volt_values_vss_relative"][1]).all()
    assert np.isnan(mat["volt_values_rail_normalized"][1]).all()
    assert int(mat["frame_avss_uv"].reshape(-1)[0]) == -2_500_000

    h5_data, _, _ = export_session_bytes(payload, "h5")
    import h5py

    with h5py.File(io.BytesIO(h5_data), "r") as handle:
        assert np.isfinite(handle["frames/volt_values_vss_relative"][0]).any()
        assert np.isnan(handle["frames/volt_values_vss_relative"][1]).all()
        assert int(handle["frames/avss_uv"][0]) == -2_500_000


def test_xlsx_contains_required_lossless_and_domain_sheets():
    data, _, _ = export_session_bytes(_payload([_session_frame("MIXED", rows=8, seq=1)]), "xlsx")
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    assert {
        "metadata", "frame_metadata", "cap_wide", "volt_wide", "res_wide",
        "quality_long", "events", "logs", "cap_offsets", "volt_vss_wide", "volt_normalized_wide",
    }.issubset(set(workbook.sheetnames))
