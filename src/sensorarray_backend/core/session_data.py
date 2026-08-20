from __future__ import annotations

import csv
import io
import json
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np

SESSION_DATA_FORMATS = {"csv", "xlsx", "mat", "h5", "zip"}
SESSION_SCHEMA_VERSION = 3
SIGNED_INT_MISSING = np.iinfo(np.int64).min

CSV_COLUMNS = [
    "schemaVersion",
    "frameIndex",
    "sessionId",
    "bootId",
    "connectionGeneration",
    "deviceTimestampUs",
    "hostTimestampUtc",
    "hostMonotonicNs",
    "seq",
    "row",
    "col",
    "cell",
    "mode",
    "unit",
    "scale",
    "rawFixed",
    "physicalValue",
    "expected",
    "expectedKnown",
    "acquired",
    "acquiredKnown",
    "valid",
    "fresh",
    "freshKnown",
    "error",
    "errorCode",
    "errorReason",
    "pga",
    "pgaBypass",
    "rowsGeneration",
    "rowsRequestId",
    "modeGeneration",
    "modeRequestId",
    "profileGeneration",
    "profileRequestId",
    "configuredRowProfile",
    "wireRowProfile",
    "avdd",
    "avss",
    "railSpan",
    "railValid",
    "railFresh",
    "railAge",
    "railSource",
    "railReason",
    "source",
    # Compatibility aliases retained for existing V2 analysis scripts.
    "measurementMode",
    "timeSeconds",
    "rows",
    "generation",
    "requestId",
    # Quantity-specific aliases keep files self-describing and preserve the
    # established CAP column for existing analysis scripts.
    "correctedPf",
    "valueUv",
    "valueV",
    "valueMilliOhm",
    "valueOhm",
    "voltageVssRelativeV",
    "voltageRailNormalised",
]


@dataclass(frozen=True)
class SessionFrame:
    seq: int
    timeSeconds: float
    rows: int
    valuesPf: np.ndarray
    valid: np.ndarray
    source: str = "history"
    measurementMode: str = "CAP"
    unit: str = "pF"
    scale: int = -6
    physicalValues: np.ndarray | None = None
    rawFixed: np.ndarray | None = None
    fresh: np.ndarray | None = None
    errorCodes: np.ndarray | None = None
    pga: np.ndarray | None = None
    generation: int | None = None
    requestId: int | None = None
    sessionId: str = ""
    connectionGeneration: int = 0
    bootId: int | None = None
    deviceTimestampUs: int | None = None
    hostReceivedUtc: str = ""
    hostWallTime: float | None = None
    hostReceivedMonotonicNs: int | None = None
    frameKind: str | None = None
    expected: np.ndarray | None = None
    acquired: np.ndarray | None = None
    expectedKnown: np.ndarray | None = None
    acquiredKnown: np.ndarray | None = None
    freshKnown: np.ndarray | None = None
    error: np.ndarray | None = None
    errorReasons: np.ndarray | None = None
    pgaBypass: np.ndarray | None = None
    rowsGeneration: int | None = None
    rowsRequestId: int | None = None
    modeGeneration: int | None = None
    modeRequestId: int | None = None
    profileGeneration: int | None = None
    profileRequestId: int | None = None
    configuredRowProfile: tuple[str, ...] | None = None
    wireRowProfile: str | None = None
    rowModes: tuple[str, ...] | None = None
    rowUnits: tuple[str, ...] | None = None
    rowScales: tuple[int | None, ...] | None = None
    rail: dict[str, Any] | None = None

    @property
    def mode(self) -> str:
        return _mode_value(self.measurementMode)

    def physical_values(self) -> np.ndarray:
        source = self.physicalValues if self.physicalValues is not None else self.valuesPf
        return np.asarray(source, dtype=np.float64).reshape(64)

    def values_matrix(self) -> np.ndarray:
        return self.physical_values().reshape(8, 8)

    def valid_matrix(self) -> np.ndarray:
        return np.asarray(self.valid, dtype=bool).reshape(8, 8)

    def fresh_values(self) -> np.ndarray:
        return (
            np.asarray(self.fresh, dtype=bool).reshape(64)
            if self.fresh is not None
            else np.zeros(64, dtype=bool)
        )

    def expected_values(self) -> np.ndarray:
        return np.asarray(self.expected, dtype=bool).reshape(64) if self.expected is not None else np.zeros(64, dtype=bool)

    def acquired_values(self) -> np.ndarray:
        return np.asarray(self.acquired, dtype=bool).reshape(64) if self.acquired is not None else np.zeros(64, dtype=bool)

    def expected_known_values(self) -> np.ndarray:
        return _known_values(self.expectedKnown, self.expected is not None)

    def acquired_known_values(self) -> np.ndarray:
        return _known_values(self.acquiredKnown, self.acquired is not None)

    def fresh_known_values(self) -> np.ndarray:
        return _known_values(self.freshKnown, self.fresh is not None)

    def error_values(self) -> np.ndarray:
        if self.error is not None:
            return np.asarray(self.error, dtype=bool).reshape(64)
        return self.error_code_values() >= 0

    def error_reason_values(self) -> np.ndarray:
        if self.errorReasons is not None:
            return np.asarray(self.errorReasons, dtype=object).reshape(64)
        return np.full(64, "", dtype=object)

    def pga_bypass_values(self) -> np.ndarray:
        if self.pgaBypass is not None:
            return np.asarray(self.pgaBypass, dtype=bool).reshape(64)
        return self.pga_values() == 0

    def row_mode_values(self) -> tuple[str, ...]:
        if self.rowModes is not None:
            return _profile_tuple(self.rowModes, allow_none=True)
        return (self.mode,) * 8

    def row_unit_values(self) -> tuple[str, ...]:
        if self.rowUnits is not None:
            return _string_tuple(self.rowUnits, self.unit)
        return tuple(_mode_unit_scale(mode)[0] if mode != "NONE" else "" for mode in self.row_mode_values())

    def row_scale_values(self) -> tuple[int | None, ...]:
        if self.rowScales is not None:
            values = tuple(None if value is None else int(value) for value in self.rowScales)
            if len(values) != 8:
                raise ValueError("rowScales must contain exactly 8 values")
            return values
        return tuple(_mode_unit_scale(mode)[1] if mode != "NONE" else None for mode in self.row_mode_values())

    def raw_fixed_values(self) -> np.ndarray:
        if self.rawFixed is not None:
            return np.asarray(self.rawFixed, dtype=np.float64).reshape(64)
        values = self.physical_values()
        factor = 10.0 ** int(self.scale)
        return np.where(np.isfinite(values), np.rint(values / factor), np.nan)

    def error_code_values(self) -> np.ndarray:
        if self.errorCodes is not None:
            return np.asarray(self.errorCodes, dtype=np.int16).reshape(64)
        return np.where(np.asarray(self.valid, dtype=bool).reshape(64), -1, 20).astype(np.int16)

    def pga_values(self) -> np.ndarray:
        return np.asarray(self.pga, dtype=np.int16).reshape(64) if self.pga is not None else np.full(64, -1, dtype=np.int16)


@dataclass(frozen=True)
class SessionModel:
    metadata: dict[str, Any]
    display: dict[str, Any]
    offsetsPf: np.ndarray
    currentMatrix: np.ndarray
    currentValidMask: np.ndarray
    frames: list[SessionFrame]
    rawLogs: list[dict[str, Any]]
    events: list[dict[str, Any]]

    @property
    def currentMatrixPf(self) -> np.ndarray:
        return self.currentMatrix


def normalise_session_format(value: str | None) -> str:
    fmt = str(value or "h5").strip().lower().lstrip(".")
    if fmt not in SESSION_DATA_FORMATS:
        raise ValueError("session export format must be csv, xlsx, mat, h5, or zip")
    return fmt


def session_model_from_payload(payload: dict[str, Any]) -> SessionModel:
    current = payload.get("currentMatrix") if isinstance(payload.get("currentMatrix"), dict) else {}
    current_values = _matrix_from_json(current.get("values") if current.get("values") is not None else current.get("correctedPf"))
    current_valid = _bool_matrix_from_json(current.get("valid") if current.get("valid") is not None else current.get("validMask"))
    offsets = _matrix_from_json(payload.get("offsetsPf"), default=0.0)
    frames = _frames_from_payload(payload, current_values, current_valid)
    return SessionModel(
        metadata={"schemaVersion": SESSION_SCHEMA_VERSION, **dict(payload.get("metadata") or {})},
        display=dict(payload.get("display") or {}),
        offsetsPf=offsets,
        currentMatrix=current_values,
        currentValidMask=current_valid,
        frames=frames,
        rawLogs=[row for row in payload.get("rawLogs", []) if isinstance(row, dict)],
        events=[row for row in payload.get("events", []) if isinstance(row, dict)],
    )


def export_session_bytes(payload: dict[str, Any], fmt: str) -> tuple[bytes, str, str]:
    selected = normalise_session_format(fmt)
    model = session_model_from_payload(payload)
    if selected == "csv":
        return _export_csv(model), "text/csv; charset=utf-8", "csv"
    if selected == "xlsx":
        return _export_xlsx(model), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    if selected == "mat":
        return _export_mat(model), "application/octet-stream", "mat"
    if selected == "zip":
        return _export_zip(model), "application/zip", "zip"
    return _export_h5(model), "application/x-hdf5", "h5"


def load_session_frames(path: str | Path) -> list[SessionFrame]:
    source = Path(path)
    suffix = source.suffix.lower().lstrip(".")
    if suffix == "csv":
        return _load_csv(source)
    if suffix == "xlsx":
        return _load_xlsx(source)
    if suffix == "mat":
        return _load_mat(source)
    if suffix in {"h5", "hdf5"}:
        return _load_h5(source)
    if suffix == "zip":
        return _load_zip(source)
    raise ValueError("session import format must be csv, xlsx, mat, h5, or zip")


def frames_to_measurement_ascii_bytes(frames: list[SessionFrame]) -> bytes:
    from sensorarray_app.constants import CAP_FIXED_SCALE, CAP_INVALID_SENTINEL, FDC_CIRCUIT_OFFSET_PF
    from sensorarray_app.protocol.crc import crc32_reflected

    out = bytearray()
    previous_mode: str | None = None
    previous_rows: int | None = None
    previous_profile: tuple[str, ...] | None = None
    for frame in frames:
        rows = max(1, min(8, int(frame.rows)))
        cells = rows * 8
        mode = frame.mode
        values = frame.physical_values()
        valid = np.asarray(frame.valid, dtype=bool).reshape(64)
        fresh = frame.fresh_values()
        expected, acquired = _wire_acquisition_masks(frame, cells)
        fresh &= acquired
        raw_fixed = frame.raw_fixed_values()
        error_codes = frame.error_code_values()
        generation = int(frame.generation if frame.generation is not None else 1)
        request_id = int(frame.requestId if frame.requestId is not None else 1)
        timestamp_us = int(
            frame.deviceTimestampUs
            if frame.deviceTimestampUs is not None
            else float(frame.timeSeconds) * 1_000_000
        )
        configured_profile = frame.configuredRowProfile or tuple(
            row_mode if row_mode != "NONE" else "CAP" for row_mode in frame.row_mode_values()
        )
        if previous_rows is not None and rows != previous_rows:
            # Geometry is independently transactional. Reconstruct the RAPP
            # boundary too, otherwise a replayed ROWS change can leave the
            # authoritative store gate at the previous geometry.
            out.extend(
                f"RCMD,id={request_id},old={previous_rows},req={rows},generation={generation},status=accepted\n".encode(
                    "ascii"
                )
            )
            out.extend(
                f"RAPP,id={request_id},seq={int(frame.seq)},old={previous_rows},new={rows},"
                f"gen={generation},status=applied\n".encode("ascii")
            )
        if previous_mode is not None and mode != previous_mode and mode != "MIXED" and previous_mode != "MIXED":
            # A homogeneous quantity change is a MODE=set-all transaction.
            # Its derived configured profile also changes, but that must not
            # be misreported as an independent ROWMODES transaction.
            out.extend(
                f"MACK,id={request_id},old={previous_mode},new={mode},state=accepted\n".encode("ascii")
            )
            out.extend(
                f"MAPP,id={request_id},gen={generation},old={previous_mode},new={mode},"
                f"seq={int(frame.seq)},state=applied,transitionUs=0\n".encode("ascii")
            )
        elif previous_profile is not None and configured_profile != previous_profile:
            profile_generation = int(frame.profileGeneration if frame.profileGeneration is not None else generation)
            profile_request_id = int(frame.profileRequestId if frame.profileRequestId is not None else request_id)
            old_wire = _configured_profile_wire(previous_profile)
            new_wire = _configured_profile_wire(configured_profile)
            out.extend(
                f"RMACK,id={profile_request_id},old={old_wire},new={new_wire},state=accepted\n".encode("ascii")
            )
            out.extend(
                f"RMAPP,id={profile_request_id},gen={profile_generation},seq={int(frame.seq)},"
                f"profile={new_wire},state=applied\n".encode("ascii")
            )
        if mode == "CAP":
            fixed: list[int] = []
            for index in range(cells):
                value = values[index]
                if not bool(valid[index]) or not np.isfinite(value):
                    fixed.append(CAP_INVALID_SENTINEL)
                elif np.isfinite(raw_fixed[index]):
                    fixed.append(int(raw_fixed[index]))
                else:
                    fixed.append(int(round((float(value) + FDC_CIRCUIT_OFFSET_PF) * CAP_FIXED_SCALE)))
            # CAP freshness is a row/device acquisition property independent
            # of whether a particular cell converted to a valid capacitance.
            # Folding validity into this mask would incorrectly make every
            # cell in a row stale after round-tripping one invalid cell.
            primary_mask = _device_row_mask(fresh, rows, 0, 4)
            secondary_mask = _device_row_mask(fresh, rows, 4, 8)
            row_mask = primary_mask | secondary_mask
            header = (
                f"C,seq={int(frame.seq)},ts={timestamp_us},rows={rows},cells={cells},gen={generation},rid={request_id},"
                f"rf={row_mask:02X},pf={primary_mask:02X},sf={secondary_mask:02X},"
                f"expected={_bit_mask(expected, cells):016X},acquired={_bit_mask(acquired, cells):016X},"
                f"bad=0/0/{int((~valid[:cells]).sum())},fmt=pf6,n={cells}\n"
            ).encode("ascii")
            body = bytearray(header)
            for line_index, start in enumerate(range(0, cells, 16)):
                body.extend(f"D{line_index},{','.join(str(value) for value in fixed[start:start + 16])}\n".encode("ascii"))
            trailer_identity = f"gen={generation},rid={request_id}"
        elif mode in {"VOLT", "RES"}:
            unit = "V" if mode == "VOLT" else "ohm"
            scale = -6 if mode == "VOLT" else -3
            format_name = "uv-x" if mode == "VOLT" else "mohm-x"
            valid_mask = _bit_mask(valid, cells)
            fresh_mask = _bit_mask(fresh, cells)
            error_mask = _bit_mask(error_codes >= 0, cells)
            reference = "AVDD_AVSS" if mode == "VOLT" else "INTREF"
            rail = frame.rail or {}
            rail_valid = int(bool(rail.get("railValid", False)))
            rail_age = _int_value(rail.get("railAge"), 0)
            avdd = _int_value(rail.get("avddUv") if rail.get("avddUv") is not None else rail.get("avdd"), 0)
            avss = _int_value(rail.get("avssUv") if rail.get("avssUv") is not None else rail.get("avss"), 0)
            mode_generation = int(frame.modeGeneration if frame.modeGeneration is not None else generation)
            mode_request_id = int(frame.modeRequestId if frame.modeRequestId is not None else request_id)
            header = (
                f"{'V' if mode == 'VOLT' else 'R'},seq={int(frame.seq)},ts={timestamp_us},rows={rows},cells={cells},"
                f"gen={mode_generation},rid={mode_request_id},mode={mode},unit={unit},scale={scale},valid={valid_mask:016X},"
                f"fresh={fresh_mask:016X},error={error_mask:016X},expected={_bit_mask(expected, cells):016X},"
                f"acquired={_bit_mask(acquired, cells):016X},ref={reference},rail={rail_valid},age={rail_age},avdd={avdd},avss={avss},"
                f"vexc=0,rref=0,dur=0,tr=0,gc=0,ov=0,aa=0,fb=0,ir=0,to=0,st=0,spi=0,fmt={format_name},"
                f"n={cells},bad={int((~valid[:cells]).sum())}\n"
            ).encode("ascii")
            body = bytearray(header)
            for line_index, start in enumerate(range(0, cells, 16)):
                tokens: list[str] = []
                for index in range(start, min(start + 16, cells)):
                    if not bool(valid[index]) or not np.isfinite(raw_fixed[index]):
                        code = int(error_codes[index]) if int(error_codes[index]) >= 0 else 20
                        tokens.append(f"X{code & 0xFF:02X}")
                    else:
                        tokens.append(str(int(raw_fixed[index])))
                body.extend(f"D{line_index},{','.join(tokens)}\n".encode("ascii"))
            pga = frame.pga_values()
            for line_index, start in enumerate(range(0, cells, 16)):
                packed = "".join(f"{(int(pga[index]) if int(pga[index]) in {0, 1, 2, 4, 8, 16, 32} else 1):02X}" for index in range(start, min(start + 16, cells)))
                body.extend(f"P{line_index},{packed}\n".encode("ascii"))
            trailer_identity = f"gen={mode_generation},rid={mode_request_id}"
        else:
            rows_generation = int(frame.rowsGeneration if frame.rowsGeneration is not None else generation)
            rows_request_id = int(frame.rowsRequestId if frame.rowsRequestId is not None else request_id)
            profile_generation = int(frame.profileGeneration if frame.profileGeneration is not None else generation)
            profile_request_id = int(frame.profileRequestId if frame.profileRequestId is not None else request_id)
            wire_profile = frame.wireRowProfile or _mixed_wire_profile(frame.row_mode_values(), rows)
            row_modes = frame.row_mode_values()
            row_units = frame.row_unit_values()
            row_scales = frame.row_scale_values()
            header = (
                f"M,seq={int(frame.seq)},ts={timestamp_us},rows={rows},cells={cells},rgen={rows_generation},"
                f"rrid={rows_request_id},pgen={profile_generation},prid={profile_request_id},profile={wire_profile},"
                f"expected={_bit_mask(expected, cells):016X},acquired={_bit_mask(acquired, cells):016X},fmt=mix1\n"
            ).encode("ascii")
            body = bytearray(header)
            errors = frame.error_values()
            for row_index in range(rows):
                start = row_index * 8
                row_mode = row_modes[row_index]
                if row_mode not in {"CAP", "VOLT", "RES"}:
                    raise ValueError(f"active Mixed row {row_index + 1} has no physical mode")
                required_unit, required_scale = _mode_unit_scale(row_mode)
                if (row_units[row_index], row_scales[row_index]) != (required_unit, required_scale):
                    raise ValueError(f"Mixed row {row_index + 1} unit/scale disagrees with mode")
                fmt = {"CAP": "pf6", "VOLT": "uv-x", "RES": "mohm-x"}[row_mode]
                row_valid = _bit_mask(valid[start : start + 8], 8)
                row_fresh = _bit_mask(fresh[start : start + 8], 8)
                row_error = _bit_mask(errors[start : start + 8], 8)
                row_expected = _bit_mask(expected[start : start + 8], 8)
                row_acquired = _bit_mask(acquired[start : start + 8], 8)
                tokens: list[str] = []
                for cell_index in range(start, start + 8):
                    if not bool(valid[cell_index]) or not np.isfinite(raw_fixed[cell_index]):
                        code = int(error_codes[cell_index]) if int(error_codes[cell_index]) >= 0 else 0x14
                        tokens.append(f"X{code & 0xFF:02X}")
                    else:
                        tokens.append(str(int(raw_fixed[cell_index])))
                body.extend(
                    (
                        f"MR,s={row_index + 1},m={row_mode},unit={required_unit},scale={required_scale},"
                        f"expected={row_expected:02X},acquired={row_acquired:02X},valid={row_valid:02X},"
                        f"fresh={row_fresh:02X},error={row_error:02X},fmt={fmt},D={','.join(tokens)}\n"
                    ).encode("ascii")
                )
            trailer_identity = (
                f"rgen={rows_generation},rrid={rows_request_id},pgen={profile_generation},prid={profile_request_id}"
            )
        crc = crc32_reflected(bytes(body))
        body.extend(f"K,seq={int(frame.seq)},{trailer_identity},crc={crc:08X}\n".encode("ascii"))
        out.extend(body)
        previous_mode = mode
        previous_rows = rows
        previous_profile = configured_profile
    return bytes(out)


def frames_to_cap_ascii_bytes(frames: list[SessionFrame]) -> bytes:
    """Compatibility name retained; modern frames are emitted as C/V/R."""

    return frames_to_measurement_ascii_bytes(frames)


def _frames_from_payload(payload: dict[str, Any], current_values: np.ndarray, current_valid: np.ndarray) -> list[SessionFrame]:
    history_frames = payload.get("historyFrames")
    frames = [
        _frame_from_history_dict(index, frame)
        for index, frame in enumerate(history_frames or [])
        if isinstance(frame, dict)
    ]
    if frames:
        return frames
    metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
    current = payload.get("currentMatrix") if isinstance(payload.get("currentMatrix"), dict) else {}
    mode = _mode_value(current.get("mode") or metadata.get("measurementMode") or "CAP")
    unit, scale = _mode_unit_scale(mode)
    return [
        SessionFrame(
            seq=_int_value(metadata.get("frameSeq"), 1),
            timeSeconds=0.0,
            rows=_rows_value(metadata.get("rows"), 8),
            valuesPf=current_values.reshape(64) if mode == "CAP" else np.full(64, np.nan),
            valid=current_valid.reshape(64),
            source="current",
            measurementMode=mode,
            unit=unit,
            scale=scale,
            physicalValues=current_values.reshape(64),
            rawFixed=_flat_numeric_matrix(current.get("rawFixed")),
            fresh=_flat_bool_matrix(current.get("fresh"), default=np.zeros(64, dtype=bool)),
            freshKnown=_flat_bool_matrix(
                current.get("freshKnown"),
                default=np.full(64, isinstance(current.get("fresh"), list), dtype=bool),
            ),
            expected=_flat_bool_matrix(current.get("expected"), default=np.zeros(64, dtype=bool)),
            acquired=_flat_bool_matrix(current.get("acquired"), default=np.zeros(64, dtype=bool)),
            expectedKnown=_flat_bool_matrix(
                current.get("expectedKnown"),
                default=np.full(64, isinstance(current.get("expected"), list), dtype=bool),
            ),
            acquiredKnown=_flat_bool_matrix(
                current.get("acquiredKnown"),
                default=np.full(64, isinstance(current.get("acquired"), list), dtype=bool),
            ),
            error=_flat_bool_matrix(current.get("error"), default=np.zeros(64, dtype=bool)),
            errorCodes=_flat_int_matrix(current.get("errorCodes"), -1),
            errorReasons=_flat_string_matrix(current.get("errorReasons"), ""),
            pga=_flat_int_matrix(current.get("pga"), -1),
            pgaBypass=_flat_bool_matrix(current.get("pgaBypass"), default=np.zeros(64, dtype=bool)),
            generation=_optional_int(current.get("generation")),
            requestId=_optional_int(current.get("requestId")),
            sessionId=str(metadata.get("sessionId") or ""),
            connectionGeneration=_int_value(current.get("connectionGeneration"), 0),
            bootId=_optional_int(current.get("bootId")),
            deviceTimestampUs=_optional_int(current.get("timestampUs")),
            frameKind=mode,
            rowModes=_optional_profile(current.get("modeByRow")),
            rowUnits=_optional_string_tuple(current.get("unitByRow")),
            rowScales=_optional_scale_tuple(current.get("scaleByRow")),
            rowsGeneration=_optional_int(current.get("rowsGeneration")),
            rowsRequestId=_optional_int(current.get("rowsRequestId")),
            profileGeneration=_optional_int(current.get("profileGeneration")),
            profileRequestId=_optional_int(current.get("profileRequestId")),
            wireRowProfile=_optional_string(current.get("wireRowProfile")),
            rail=dict(payload.get("rail") or {}),
        )
    ]


def _frame_from_history_dict(index: int, frame: dict[str, Any]) -> SessionFrame:
    mode = _mode_value(frame.get("frameKind") or frame.get("measurementMode") or frame.get("mode") or "CAP")
    unit, default_scale = _mode_unit_scale(mode)
    raw_values = frame.get("physicalValues")
    if raw_values is None:
        raw_values = frame.get("values")
    if raw_values is None:
        raw_values = frame.get("valuesPf")
    values = _flat_numeric_list(raw_values)
    valid = _flat_bool_list(frame.get("valid"), default=np.isfinite(values))
    fresh_present = isinstance(frame.get("fresh"), (list, tuple, np.ndarray))
    expected_present = isinstance(frame.get("expected"), (list, tuple, np.ndarray))
    acquired_present = isinstance(frame.get("acquired"), (list, tuple, np.ndarray))
    generation = _optional_int(frame.get("generation"))
    request_id = _optional_int(frame.get("requestId"))
    timestamp_us = _optional_int(frame.get("deviceTimestampUs"))
    time_seconds = _float_value(
        frame.get("timeSeconds"),
        float(timestamp_us) / 1_000_000.0 if timestamp_us is not None else float(index),
    )
    rail = frame.get("rail") if isinstance(frame.get("rail"), dict) else {
        key: frame.get(key)
        for key in (
            "railValid", "railFresh", "railAge", "avddUv", "avssUv", "railSpanUv",
            "railSource", "railReason",
        )
        if key in frame
    }
    return SessionFrame(
        seq=_int_value(frame.get("seq"), index + 1),
        timeSeconds=time_seconds,
        rows=_rows_value(frame.get("rows"), 8),
        valuesPf=values if mode == "CAP" else np.full(64, np.nan),
        valid=valid,
        source=str(frame.get("source") or "history"),
        measurementMode=mode,
        unit=str(frame.get("unit") or unit),
        scale=_int_value(frame.get("scale"), default_scale),
        physicalValues=values,
        rawFixed=_flat_numeric_list(frame.get("rawFixed")),
        fresh=_flat_bool_list(frame.get("fresh"), default=np.zeros(64, dtype=bool)),
        expected=_flat_bool_list(frame.get("expected"), default=np.zeros(64, dtype=bool)),
        acquired=_flat_bool_list(frame.get("acquired"), default=np.zeros(64, dtype=bool)),
        expectedKnown=_flat_bool_list(
            frame.get("expectedKnown"), default=np.full(64, expected_present, dtype=bool)
        ),
        acquiredKnown=_flat_bool_list(
            frame.get("acquiredKnown"), default=np.full(64, acquired_present, dtype=bool)
        ),
        freshKnown=_flat_bool_list(
            frame.get("freshKnown"), default=np.full(64, fresh_present, dtype=bool)
        ),
        error=_flat_bool_list(frame.get("error"), default=np.zeros(64, dtype=bool)),
        errorCodes=_flat_int_list(frame.get("errorCodes"), -1),
        errorReasons=_flat_string_list(frame.get("errorReasons"), ""),
        pga=_flat_int_list(frame.get("pga"), -1),
        pgaBypass=_flat_bool_list(frame.get("pgaBypass"), default=np.zeros(64, dtype=bool)),
        generation=generation,
        requestId=request_id,
        sessionId=str(frame.get("sessionId") or ""),
        connectionGeneration=_int_value(frame.get("connectionGeneration"), 0),
        bootId=_optional_int(frame.get("bootId")),
        deviceTimestampUs=timestamp_us,
        hostReceivedUtc=str(frame.get("hostReceivedUtc") or frame.get("hostTimestampUtc") or ""),
        hostWallTime=_optional_float(frame.get("hostWallTime")),
        hostReceivedMonotonicNs=_optional_int(
            frame.get("hostReceivedMonotonicNs")
            if frame.get("hostReceivedMonotonicNs") is not None
            else frame.get("hostMonotonicNs")
        ),
        frameKind=mode,
        rowsGeneration=_optional_int(frame.get("rowsGeneration")) if mode == "MIXED" or frame.get("rowsGeneration") is not None else (generation if mode == "CAP" else None),
        rowsRequestId=_optional_int(frame.get("rowsRequestId")) if mode == "MIXED" or frame.get("rowsRequestId") is not None else (request_id if mode == "CAP" else None),
        modeGeneration=_optional_int(frame.get("modeGeneration")) if frame.get("modeGeneration") is not None else (generation if mode in {"VOLT", "RES"} else None),
        modeRequestId=_optional_int(frame.get("modeRequestId")) if frame.get("modeRequestId") is not None else (request_id if mode in {"VOLT", "RES"} else None),
        profileGeneration=_optional_int(frame.get("profileGeneration")),
        profileRequestId=_optional_int(frame.get("profileRequestId")),
        configuredRowProfile=_optional_profile(frame.get("configuredRowProfile"), allow_none=False),
        wireRowProfile=_optional_string(frame.get("wireRowProfile")),
        rowModes=_optional_profile(frame.get("rowModes")),
        rowUnits=_optional_string_tuple(frame.get("rowUnits")),
        rowScales=_optional_scale_tuple(frame.get("rowScales")),
        rail=dict(rail),
    )


def _export_csv(model: SessionModel) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    writer.writerows(_iter_frame_rows(model.frames))
    return output.getvalue().encode("utf-8")


def _export_xlsx(model: SessionModel) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc
    workbook = Workbook()
    metadata_sheet = workbook.active
    metadata_sheet.title = "metadata"
    metadata_sheet.append(["key", "value"])
    for key, value in model.metadata.items():
        metadata_sheet.append([key, _string_value(value)])
    frame_metadata = workbook.create_sheet("frame_metadata")
    metadata_columns = [
        "frameIndex", "sessionId", "bootId", "connectionGeneration", "deviceTimestampUs",
        "hostTimestampUtc", "hostMonotonicNs", "seq", "rows", "frameKind", "rowsGeneration",
        "rowsRequestId", "modeGeneration", "modeRequestId", "profileGeneration", "profileRequestId",
        "configuredRowProfile", "wireRowProfile", "railValid", "railFresh", "railAge", "avdd",
        "avss", "railSpan", "railSource", "railReason", "source",
    ]
    frame_metadata.append(metadata_columns)
    for frame_index, frame in enumerate(model.frames):
        row = _frame_metadata_row(frame_index, frame)
        frame_metadata.append([row[column] for column in metadata_columns])
    for mode, sheet_name in (("CAP", "cap_wide"), ("VOLT", "volt_wide"), ("RES", "res_wide")):
        _append_wide_sheet(workbook.create_sheet(sheet_name), model.frames, mode)
    _append_wide_sheet(workbook.create_sheet("volt_vss_wide"), model.frames, "VOLT", value_kind="vss_relative")
    _append_wide_sheet(workbook.create_sheet("volt_normalized_wide"), model.frames, "VOLT", value_kind="rail_normalized")
    canonical_rows = _iter_frame_rows(model.frames)
    quality_sheet = workbook.create_sheet("quality_long")
    quality_sheet.append(CSV_COLUMNS)
    for row in canonical_rows:
        quality_sheet.append([row[column] for column in CSV_COLUMNS])
    # Compatibility sheet retained for V2 importers; it now contains the same
    # lossless V3 mother table as quality_long.
    frames_sheet = workbook.create_sheet("frames")
    frames_sheet.append(CSV_COLUMNS)
    for row in canonical_rows:
        frames_sheet.append([row[column] for column in CSV_COLUMNS])
    offsets_sheet = workbook.create_sheet("cap_offsets")
    offsets_sheet.append(["row", "col", "offsetPf"])
    for row_index in range(8):
        for col_index in range(8):
            offsets_sheet.append([row_index + 1, col_index + 1, float(model.offsetsPf[row_index, col_index])])
    events_sheet = workbook.create_sheet("events")
    _append_mapping_sheet(events_sheet, model.events)
    logs_sheet = workbook.create_sheet("logs")
    logs_sheet.append(["timestamp", "source", "channel", "tag", "category", "severity", "rawText"])
    for row in model.rawLogs:
        logs_sheet.append([row.get("timestamp"), row.get("source"), row.get("channel"), row.get("tag"), row.get("category"), row.get("severity"), row.get("rawText")])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _export_mat(model: SessionModel) -> bytes:
    try:
        from scipy.io import savemat
    except ImportError as exc:
        raise RuntimeError("scipy is required for MAT export") from exc
    arrays = _frame_arrays(model.frames)
    payload: dict[str, Any] = {
        "session_schema_version": np.asarray([SESSION_SCHEMA_VERSION], dtype=np.int16),
        "current_matrix": model.currentMatrix,
        "current_valid_mask": model.currentValidMask.astype(np.uint8),
        "offsets_pf": model.offsetsPf,
        "events_json": np.asarray([_json_text(event) for event in model.events], dtype=object),
        "cap_values": arrays["frames_cap_values"],
        "volt_values_ground": arrays["frames_volt_values_ground"],
        "volt_values_vss_relative": arrays["frames_volt_values_vss_relative"],
        "volt_values_rail_normalized": arrays["frames_volt_values_rail_normalized"],
        "res_values": arrays["frames_res_values"],
        "row_modes": arrays["frames_row_modes"],
        "wire_profiles": arrays["frame_wire_profiles"],
        "expected_mask": arrays["frames_expected_mask"],
        "acquired_mask": arrays["frames_acquired_mask"],
        "fresh_mask": arrays["frames_fresh_mask"],
        "valid_mask": arrays["frames_valid_mask"],
        "error_codes": arrays["frames_error_codes"],
        "pga": arrays["frames_pga"],
        **arrays,
    }
    if all(frame.mode == "CAP" for frame in model.frames):
        payload["current_matrix_pf"] = model.currentMatrix
        payload["frames_values_pf"] = arrays["frames_physical_values"]
    output = io.BytesIO()
    savemat(output, payload)
    return output.getvalue()


def _export_h5(model: SessionModel) -> bytes:
    try:
        import h5py
    except ImportError as exc:
        raise RuntimeError("h5py is required for HDF5 export") from exc
    arrays = _frame_arrays(model.frames)
    output = io.BytesIO()
    with h5py.File(output, "w") as handle:
        metadata = handle.create_group("metadata")
        metadata.attrs["schema_version"] = SESSION_SCHEMA_VERSION
        for key, value in model.metadata.items():
            metadata.attrs[str(key)] = _string_value(value)
        current = handle.create_group("current")
        current.create_dataset("matrix", data=model.currentMatrix)
        current.create_dataset("valid_mask", data=model.currentValidMask.astype(np.uint8))
        frames = handle.create_group("frames")
        for name, value in arrays.items():
            if value.dtype.kind in {"U", "O"}:
                dataset_name = name.removeprefix("frame_")
                dataset_name = dataset_name.removeprefix("frames_")
                frames.create_dataset(dataset_name, data=np.asarray(value, dtype="S128"))
            elif name.startswith("frame_"):
                frames.create_dataset(name.removeprefix("frame_"), data=value)
            elif name.startswith("frames_"):
                frames.create_dataset(name.removeprefix("frames_"), data=value)
        offsets = handle.create_group("capacitance")
        offsets.create_dataset("offsets_pf", data=model.offsetsPf)
        if all(frame.mode == "CAP" for frame in model.frames):
            current.create_dataset("matrix_pf", data=model.currentMatrix)
            frames.create_dataset("values_pf", data=arrays["frames_physical_values"])
        string_dtype = h5py.string_dtype(encoding="utf-8")
        handle.create_dataset("events", data=np.asarray([_json_text(event) for event in model.events], dtype=object), dtype=string_dtype)
    return output.getvalue()


def _export_zip(model: SessionModel) -> bytes:
    """Return the lossless multi-table CSV bundle required for Mixed data."""

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("metadata.csv", _key_value_csv(model.metadata))
        archive.writestr("frame_metadata.csv", _frame_metadata_csv(model.frames))
        archive.writestr("cap_wide.csv", _wide_csv(model.frames, "CAP"))
        archive.writestr("volt_wide.csv", _wide_csv(model.frames, "VOLT"))
        archive.writestr("res_wide.csv", _wide_csv(model.frames, "RES"))
        archive.writestr("volt_vss_wide.csv", _wide_csv(model.frames, "VOLT", value_kind="vss_relative"))
        archive.writestr("volt_normalized_wide.csv", _wide_csv(model.frames, "VOLT", value_kind="rail_normalized"))
        archive.writestr("quality_long.csv", _export_csv(model))
        archive.writestr("events.csv", _mapping_csv(model.events, ("kind", "hostTimestampUtc")))
        archive.writestr(
            "logs.csv",
            _mapping_csv(
                model.rawLogs,
                ("timestamp", "source", "channel", "tag", "category", "severity", "rawText"),
            ),
        )
        archive.writestr("cap_offsets.csv", _offsets_csv(model.offsetsPf))
    return output.getvalue()


def _load_csv(path: Path) -> list[SessionFrame]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = set(reader.fieldnames or [])
        required = {"frameIndex", "seq", "rows", "row", "col", "valid"}
        missing = sorted(required - fields)
        if missing or not ({"physicalValue", "correctedPf"} & fields):
            detail = missing or ["physicalValue or correctedPf"]
            raise ValueError(f"CSV missing required columns: {', '.join(detail)}")
        return _rows_to_frames(reader)


def _load_xlsx(path: Path) -> list[SessionFrame]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX import") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "frames" not in workbook.sheetnames:
        raise ValueError("XLSX missing required sheet: frames")
    rows = list(workbook["frames"].iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX frames sheet is empty")
    header = [str(value or "") for value in rows[0]]
    return _rows_to_frames(dict(zip(header, row, strict=False)) for row in rows[1:])


def _load_mat(path: Path) -> list[SessionFrame]:
    try:
        from scipy.io import loadmat
    except ImportError as exc:
        raise RuntimeError("scipy is required for MAT import") from exc
    payload = loadmat(path)
    if "frames_physical_values" in payload:
        return _arrays_to_frames(
            payload["frames_physical_values"], payload["frames_valid_mask"], payload.get("frames_fresh_mask"),
            payload.get("frames_raw_fixed"), payload.get("frames_error_codes"), payload.get("frames_pga"),
            payload["frame_seq"].ravel(), payload["frame_time_seconds"].ravel(), payload["frame_rows"].ravel(),
            payload.get("frame_mode"), payload.get("frame_unit"), payload.get("frame_scale"), payload.get("frame_source"),
            payload.get("frame_generation"), payload.get("frame_request_id"),
            payload,
        )
    required = ["frames_values_pf", "frames_valid_mask", "frame_seq", "frame_time_seconds", "frame_rows"]
    missing = [name for name in required if name not in payload]
    if missing:
        raise ValueError(f"MAT missing required variables: {', '.join(missing)}")
    return _arrays_to_frames(
        payload["frames_values_pf"], payload["frames_valid_mask"], None, None, None, None,
        payload["frame_seq"].ravel(), payload["frame_time_seconds"].ravel(), payload["frame_rows"].ravel(),
        None, None, None, None, None, None,
    )


def _load_h5(path: Path) -> list[SessionFrame]:
    try:
        import h5py
    except ImportError as exc:
        raise RuntimeError("h5py is required for HDF5 import") from exc
    with h5py.File(path, "r") as handle:
        if "frames/physical_values" in handle:
            return _arrays_to_frames(
                handle["frames/physical_values"][()], handle["frames/valid_mask"][()],
                _h5_optional(handle, "frames/fresh_mask"), _h5_optional(handle, "frames/raw_fixed"),
                _h5_optional(handle, "frames/error_codes"), _h5_optional(handle, "frames/pga"),
                handle["frames/seq"][()], handle["frames/time_seconds"][()], handle["frames/rows"][()],
                _h5_optional(handle, "frames/mode"), _h5_optional(handle, "frames/unit"),
                _h5_optional(handle, "frames/scale"), _h5_optional(handle, "frames/source"),
                _h5_optional(handle, "frames/generation"), _h5_optional(handle, "frames/request_id"),
                _h5_group_dict(handle, "frames"),
            )
        required = ["frames/values_pf", "frames/valid_mask", "frames/seq", "frames/time_seconds", "frames/rows"]
        missing = [name for name in required if name not in handle]
        if missing:
            raise ValueError(f"HDF5 missing required datasets: {', '.join(missing)}")
        return _arrays_to_frames(
            handle["frames/values_pf"][()], handle["frames/valid_mask"][()], None, None, None, None,
            handle["frames/seq"][()], handle["frames/time_seconds"][()], handle["frames/rows"][()],
            None, None, None, None, None, None,
        )


def _load_zip(path: Path) -> list[SessionFrame]:
    try:
        with zipfile.ZipFile(path, "r") as archive:
            if "quality_long.csv" not in archive.namelist():
                raise ValueError("ZIP session bundle is missing quality_long.csv")
            with archive.open("quality_long.csv", "r") as raw:
                text = io.TextIOWrapper(raw, encoding="utf-8", newline="")
                reader = csv.DictReader(text)
                return _rows_to_frames(reader)
    except zipfile.BadZipFile as exc:
        raise ValueError("invalid ZIP session bundle") from exc


def _frame_arrays(frames: list[SessionFrame]) -> dict[str, np.ndarray]:
    physical = np.stack([frame.values_matrix() for frame in frames], axis=0)
    row_modes = np.asarray([frame.row_mode_values() for frame in frames], dtype="U8")
    cap_values = np.where(row_modes[:, :, None] == "CAP", physical, np.nan)
    volt_values = np.where(row_modes[:, :, None] == "VOLT", physical, np.nan)
    res_values = np.where(row_modes[:, :, None] == "RES", physical, np.nan)
    derived = [_derived_voltage_values(frame) for frame in frames]
    volt_vss_relative = np.stack([item[0].reshape(8, 8) for item in derived], axis=0)
    volt_rail_normalized = np.stack([item[1].reshape(8, 8) for item in derived], axis=0)
    rail_valid = np.asarray([bool((frame.rail or {}).get("railValid", False)) for frame in frames], dtype=np.uint8)
    rail_fresh = np.asarray([bool((frame.rail or {}).get("railFresh", False)) for frame in frames], dtype=np.uint8)
    return {
        "frames_physical_values": physical,
        "frames_cap_values": cap_values,
        "frames_volt_values_ground": volt_values,
        "frames_volt_values_vss_relative": volt_vss_relative,
        "frames_volt_values_rail_normalized": volt_rail_normalized,
        "frames_res_values": res_values,
        "frames_raw_fixed": np.stack([frame.raw_fixed_values().reshape(8, 8) for frame in frames], axis=0),
        "frames_valid_mask": np.stack([frame.valid_matrix().astype(np.uint8) for frame in frames], axis=0),
        "frames_fresh_mask": np.stack([frame.fresh_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_expected_mask": np.stack([frame.expected_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_acquired_mask": np.stack([frame.acquired_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_expected_known_mask": np.stack([frame.expected_known_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_acquired_known_mask": np.stack([frame.acquired_known_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_fresh_known_mask": np.stack([frame.fresh_known_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_error_mask": np.stack([frame.error_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frames_error_codes": np.stack([frame.error_code_values().reshape(8, 8) for frame in frames], axis=0),
        "frames_error_reasons": np.stack([frame.error_reason_values().reshape(8, 8) for frame in frames], axis=0).astype("U96"),
        "frames_pga": np.stack([frame.pga_values().reshape(8, 8) for frame in frames], axis=0),
        "frames_pga_bypass": np.stack([frame.pga_bypass_values().reshape(8, 8).astype(np.uint8) for frame in frames], axis=0),
        "frame_seq": np.asarray([frame.seq for frame in frames], dtype=np.int64),
        "frame_time_seconds": np.asarray([frame.timeSeconds for frame in frames], dtype=np.float64),
        "frame_device_timestamp_us": np.asarray([_array_int_or_missing(frame.deviceTimestampUs) for frame in frames], dtype=np.int64),
        "frame_host_timestamp_utc": np.asarray([frame.hostReceivedUtc for frame in frames], dtype="U64"),
        "frame_host_monotonic_ns": np.asarray([_array_int_or_missing(frame.hostReceivedMonotonicNs) for frame in frames], dtype=np.int64),
        "frame_rows": np.asarray([frame.rows for frame in frames], dtype=np.int16),
        "frame_mode": np.asarray([frame.mode for frame in frames], dtype="U8"),
        "frame_unit": np.asarray([frame.unit for frame in frames], dtype="U8"),
        "frame_scale": np.asarray([frame.scale for frame in frames], dtype=np.int16),
        "frame_source": np.asarray([frame.source for frame in frames], dtype="U32"),
        "frame_generation": np.asarray([-1 if frame.generation is None else frame.generation for frame in frames], dtype=np.int64),
        "frame_request_id": np.asarray([-1 if frame.requestId is None else frame.requestId for frame in frames], dtype=np.int64),
        "frame_session_id": np.asarray([frame.sessionId for frame in frames], dtype="U64"),
        "frame_connection_generation": np.asarray([frame.connectionGeneration for frame in frames], dtype=np.int64),
        "frame_boot_id": np.asarray([_array_int_or_missing(frame.bootId) for frame in frames], dtype=np.int64),
        "frame_rows_generation": np.asarray([_array_int_or_missing(frame.rowsGeneration) for frame in frames], dtype=np.int64),
        "frame_rows_request_id": np.asarray([_array_int_or_missing(frame.rowsRequestId) for frame in frames], dtype=np.int64),
        "frame_mode_generation": np.asarray([_array_int_or_missing(frame.modeGeneration) for frame in frames], dtype=np.int64),
        "frame_mode_request_id": np.asarray([_array_int_or_missing(frame.modeRequestId) for frame in frames], dtype=np.int64),
        "frame_profile_generation": np.asarray([_array_int_or_missing(frame.profileGeneration) for frame in frames], dtype=np.int64),
        "frame_profile_request_id": np.asarray([_array_int_or_missing(frame.profileRequestId) for frame in frames], dtype=np.int64),
        "frames_row_modes": row_modes,
        "frames_row_units": np.asarray([frame.row_unit_values() for frame in frames], dtype="U8"),
        "frames_row_scales": np.asarray(
            [[-128 if value is None else value for value in frame.row_scale_values()] for frame in frames], dtype=np.int16
        ),
        "frame_wire_profiles": np.asarray([frame.wireRowProfile or "" for frame in frames], dtype="U8"),
        "frame_configured_profiles": np.asarray([_profile_text(frame.configuredRowProfile) for frame in frames], dtype="U8"),
        "frame_rail_valid": rail_valid,
        "frame_rail_fresh": rail_fresh,
        "frame_rail_age": np.asarray([_array_int_or_missing((frame.rail or {}).get("railAge")) for frame in frames], dtype=np.int64),
        "frame_avdd_uv": np.asarray([_array_int_or_missing(_rail_value(frame.rail or {}, "avddUv", "avdd")) for frame in frames], dtype=np.int64),
        "frame_avss_uv": np.asarray([_array_signed_int_or_missing(_rail_value(frame.rail or {}, "avssUv", "avss")) for frame in frames], dtype=np.int64),
        "frame_rail_span_uv": np.asarray([_array_int_or_missing(_rail_value(frame.rail or {}, "railSpanUv", "railSpan")) for frame in frames], dtype=np.int64),
        "frame_rail_boot_id": np.asarray([_array_int_or_missing((frame.rail or {}).get("bootId")) for frame in frames], dtype=np.int64),
    }


def _arrays_to_frames(
    values: Any, valid: Any, fresh: Any, raw_fixed: Any, error_codes: Any, pga: Any,
    seq: Any, time_seconds: Any, rows: Any, modes: Any, units: Any, scales: Any, sources: Any,
    generations: Any, request_ids: Any, extras: dict[str, Any] | None = None,
) -> list[SessionFrame]:
    values_array = np.asarray(values, dtype=np.float64)
    valid_array = np.asarray(valid, dtype=bool)
    if values_array.ndim != 3 or values_array.shape[1:] != (8, 8):
        raise ValueError("frames physical values must have shape Nx8x8")
    if valid_array.shape != values_array.shape:
        raise ValueError("frames valid mask must match physical values")
    count = values_array.shape[0]
    frames: list[SessionFrame] = []
    for index in range(count):
        mode = _array_mode(modes, index)
        unit, default_scale = _mode_unit_scale(mode)
        generation = _array_optional_int(generations, index)
        request_id = _array_optional_int(request_ids, index)
        extra = extras or {}
        expected_array = _extra_value(extra, "frames_expected_mask", "expected_mask")
        acquired_array = _extra_value(extra, "frames_acquired_mask", "acquired_mask")
        fresh_known_array = _extra_value(extra, "frames_fresh_known_mask", "fresh_known_mask")
        expected_known_array = _extra_value(extra, "frames_expected_known_mask", "expected_known_mask")
        acquired_known_array = _extra_value(extra, "frames_acquired_known_mask", "acquired_known_mask")
        row_modes_array = _extra_value(extra, "frames_row_modes", "row_modes")
        parsed_row_modes = _array_row_strings(row_modes_array, index)
        frames.append(
            SessionFrame(
                seq=_int_value(seq[index] if index < len(seq) else index + 1, index + 1),
                timeSeconds=_float_value(time_seconds[index] if index < len(time_seconds) else index, float(index)),
                rows=_rows_value(rows[index] if index < len(rows) else 8, 8),
                valuesPf=values_array[index].reshape(64) if mode == "CAP" else np.full(64, np.nan),
                valid=valid_array[index].reshape(64),
                measurementMode=mode,
                unit=_array_string(units, index, unit),
                scale=_array_int(scales, index, default_scale),
                physicalValues=values_array[index].reshape(64),
                rawFixed=_array_matrix(raw_fixed, index, np.nan),
                fresh=_array_matrix(fresh, index, False).astype(bool),
                freshKnown=_array_matrix(fresh_known_array, index, fresh is not None).astype(bool),
                expected=_array_matrix(expected_array, index, False).astype(bool),
                acquired=_array_matrix(acquired_array, index, False).astype(bool),
                expectedKnown=_array_matrix(expected_known_array, index, expected_array is not None).astype(bool),
                acquiredKnown=_array_matrix(acquired_known_array, index, acquired_array is not None).astype(bool),
                error=_array_matrix(_extra_value(extra, "frames_error_mask", "error_mask"), index, False).astype(bool),
                errorCodes=_array_matrix(error_codes, index, -1).astype(np.int16),
                errorReasons=_array_matrix(_extra_value(extra, "frames_error_reasons", "error_reasons"), index, "").astype(object),
                pga=_array_matrix(pga, index, -1).astype(np.int16),
                pgaBypass=_array_matrix(_extra_value(extra, "frames_pga_bypass", "pga_bypass"), index, False).astype(bool),
                source=_array_string(sources, index, "history"),
                generation=generation,
                requestId=request_id,
                sessionId=_array_string(_extra_value(extra, "frame_session_id", "session_id"), index, ""),
                connectionGeneration=_array_int(_extra_value(extra, "frame_connection_generation", "connection_generation"), index, 0),
                bootId=_array_optional_int(_extra_value(extra, "frame_boot_id", "boot_id"), index),
                deviceTimestampUs=_array_optional_int(
                    _extra_value(extra, "frame_device_timestamp_us", "device_timestamp_us"), index
                ),
                hostReceivedUtc=_array_string(
                    _extra_value(extra, "frame_host_timestamp_utc", "host_timestamp_utc"), index, ""
                ),
                hostReceivedMonotonicNs=_array_optional_int(
                    _extra_value(extra, "frame_host_monotonic_ns", "host_monotonic_ns"), index
                ),
                frameKind=mode,
                rowsGeneration=_array_optional_int(_extra_value(extra, "frame_rows_generation", "rows_generation"), index),
                rowsRequestId=_array_optional_int(_extra_value(extra, "frame_rows_request_id", "rows_request_id"), index),
                modeGeneration=_array_optional_int(_extra_value(extra, "frame_mode_generation", "mode_generation"), index),
                modeRequestId=_array_optional_int(_extra_value(extra, "frame_mode_request_id", "mode_request_id"), index),
                profileGeneration=_array_optional_int(
                    _extra_value(extra, "frame_profile_generation", "profile_generation"), index
                ),
                profileRequestId=_array_optional_int(
                    _extra_value(extra, "frame_profile_request_id", "profile_request_id"), index
                ),
                configuredRowProfile=_array_profile(
                    _extra_value(extra, "frame_configured_profiles", "configured_profiles"), index, allow_none=False
                ),
                wireRowProfile=_array_optional_string(
                    _extra_value(extra, "frame_wire_profiles", "wire_profiles"), index
                ),
                rowModes=parsed_row_modes,
                rowUnits=_array_row_strings(_extra_value(extra, "frames_row_units", "row_units"), index),
                rowScales=_array_row_scales(_extra_value(extra, "frames_row_scales", "row_scales"), index),
                rail={
                    "railValid": bool(_array_int(_extra_value(extra, "frame_rail_valid", "rail_valid"), index, 0)),
                    "railFresh": bool(_array_int(_extra_value(extra, "frame_rail_fresh", "rail_fresh"), index, 0)),
                    "railAge": _array_optional_int(_extra_value(extra, "frame_rail_age", "rail_age"), index),
                    "avddUv": _array_optional_int(_extra_value(extra, "frame_avdd_uv", "avdd_uv"), index),
                    "avssUv": _array_optional_signed_int(_extra_value(extra, "frame_avss_uv", "avss_uv"), index),
                    "railSpanUv": _array_optional_int(
                        _extra_value(extra, "frame_rail_span_uv", "rail_span_uv"), index
                    ),
                    "bootId": _array_optional_int(_extra_value(extra, "frame_rail_boot_id", "rail_boot_id"), index),
                },
            )
        )
    if not frames:
        raise ValueError("session file contains no frames")
    return frames


def _rows_to_frames(rows: Iterable[dict[str, Any]]) -> list[SessionFrame]:
    groups: dict[int, dict[str, Any]] = {}
    for row in rows:
        frame_index = _int_value(row.get("frameIndex"), 0)
        frame_mode = _mode_value(row.get("measurementMode") or row.get("frameKind") or row.get("mode") or "CAP")
        cell_mode = str(row.get("mode") or (frame_mode if frame_mode != "MIXED" else "NONE")).strip().upper()
        cell_mode = {"C": "CAP", "V": "VOLT", "R": "RES", "N": "NONE"}.get(cell_mode, cell_mode)
        if cell_mode not in {"CAP", "VOLT", "RES", "NONE"}:
            raise ValueError(f"unsupported cell mode: {cell_mode}")
        unit, default_scale = _mode_unit_scale(cell_mode if cell_mode != "NONE" else "MIXED")
        group = groups.setdefault(
            frame_index,
            {
                "seq": _int_value(row.get("seq"), frame_index + 1), "time": _float_value(row.get("timeSeconds"), float(frame_index)),
                "rows": _rows_value(row.get("rows"), 8), "mode": frame_mode, "unit": str(row.get("unit") or unit),
                "scale": _int_value(row.get("scale"), default_scale), "source": str(row.get("source") or "import"),
                "values": np.full(64, np.nan), "raw": np.full(64, np.nan), "valid": np.zeros(64, dtype=bool),
                "fresh": np.zeros(64, dtype=bool), "freshKnown": np.zeros(64, dtype=bool),
                "expected": np.zeros(64, dtype=bool), "expectedKnown": np.zeros(64, dtype=bool),
                "acquired": np.zeros(64, dtype=bool), "acquiredKnown": np.zeros(64, dtype=bool),
                "errorMask": np.zeros(64, dtype=bool), "error": np.full(64, -1, dtype=np.int16),
                "errorReasons": np.full(64, "", dtype=object), "pga": np.full(64, -1, dtype=np.int16),
                "pgaBypass": np.zeros(64, dtype=bool),
                "generation": _optional_int(row.get("generation")), "requestId": _optional_int(row.get("requestId")),
                "rowModes": ["NONE"] * 8, "rowUnits": [""] * 8, "rowScales": [None] * 8,
                "sessionId": str(row.get("sessionId") or ""),
                "bootId": _optional_int(row.get("bootId")),
                "connectionGeneration": _int_value(row.get("connectionGeneration"), 0),
                "deviceTimestampUs": _optional_int(row.get("deviceTimestampUs")),
                "hostTimestampUtc": str(row.get("hostTimestampUtc") or ""),
                "hostMonotonicNs": _optional_int(row.get("hostMonotonicNs")),
                "rowsGeneration": _optional_int(row.get("rowsGeneration")),
                "rowsRequestId": _optional_int(row.get("rowsRequestId")),
                "modeGeneration": _optional_int(row.get("modeGeneration")),
                "modeRequestId": _optional_int(row.get("modeRequestId")),
                "profileGeneration": _optional_int(row.get("profileGeneration")),
                "profileRequestId": _optional_int(row.get("profileRequestId")),
                "configuredRowProfile": _optional_profile(row.get("configuredRowProfile"), allow_none=False),
                "wireRowProfile": _optional_string(row.get("wireRowProfile")),
                "rail": {
                    "avddUv": _optional_int(row.get("avdd")), "avssUv": _optional_int(row.get("avss")),
                    "railSpanUv": _optional_int(row.get("railSpan")), "railValid": _bool_value(row.get("railValid")),
                    "railFresh": _bool_value(row.get("railFresh")), "railAge": _optional_int(row.get("railAge")),
                    "railSource": str(row.get("railSource") or ""), "railReason": str(row.get("railReason") or ""),
                },
            },
        )
        row_index, col_index = _row_index(row.get("row")), _col_index(row.get("col"))
        cell_index = row_index * 8 + col_index
        group["rowModes"][row_index] = cell_mode
        group["rowUnits"][row_index] = str(row.get("unit") or unit)
        group["rowScales"][row_index] = (
            _optional_int(row.get("scale")) if row.get("scale") not in {None, ""} else (default_scale if cell_mode != "NONE" else None)
        )
        physical = _json_number(row.get("physicalValue"))
        if physical is None:
            physical = _json_number(row.get("correctedPf"))
        if physical is not None:
            group["values"][cell_index] = physical
        raw = _json_number(row.get("rawFixed"))
        if raw is not None:
            group["raw"][cell_index] = raw
        group["valid"][cell_index] = _bool_value(row.get("valid"))
        if row.get("fresh") not in {None, ""}:
            group["fresh"][cell_index] = _bool_value(row.get("fresh"))
            group["freshKnown"][cell_index] = _bool_value(row.get("freshKnown")) if row.get("freshKnown") not in {None, ""} else True
        if row.get("expected") not in {None, ""}:
            group["expected"][cell_index] = _bool_value(row.get("expected"))
            group["expectedKnown"][cell_index] = _bool_value(row.get("expectedKnown")) if row.get("expectedKnown") not in {None, ""} else True
        if row.get("acquired") not in {None, ""}:
            group["acquired"][cell_index] = _bool_value(row.get("acquired"))
            group["acquiredKnown"][cell_index] = _bool_value(row.get("acquiredKnown")) if row.get("acquiredKnown") not in {None, ""} else True
        group["errorMask"][cell_index] = _bool_value(row.get("error"))
        group["error"][cell_index] = _int_value(row.get("errorCode"), -1)
        group["errorReasons"][cell_index] = str(row.get("errorReason") or "")
        group["pga"][cell_index] = _int_value(row.get("pga"), -1)
        group["pgaBypass"][cell_index] = _bool_value(row.get("pgaBypass"))
    frames: list[SessionFrame] = []
    for _, group in sorted(groups.items()):
        active_modes = tuple(group["rowModes"][: group["rows"]])
        inferred_mode = active_modes[0] if active_modes and len(set(active_modes)) == 1 else "MIXED"
        mode = group["mode"] if group["mode"] == "MIXED" else inferred_mode
        if mode == "NONE":
            raise ValueError("session frame has no active physical row mode")
        unit, scale = _mode_unit_scale(mode)
        frames.append(
            SessionFrame(
                seq=group["seq"], timeSeconds=group["time"], rows=group["rows"],
                valuesPf=group["values"] if mode == "CAP" else np.full(64, np.nan), valid=group["valid"], source=group["source"],
                measurementMode=mode, unit=unit, scale=scale, physicalValues=group["values"], rawFixed=group["raw"],
                fresh=group["fresh"], freshKnown=group["freshKnown"], expected=group["expected"],
                expectedKnown=group["expectedKnown"], acquired=group["acquired"], acquiredKnown=group["acquiredKnown"],
                error=group["errorMask"], errorCodes=group["error"], errorReasons=group["errorReasons"],
                pga=group["pga"], pgaBypass=group["pgaBypass"], generation=group["generation"], requestId=group["requestId"],
                sessionId=group["sessionId"], bootId=group["bootId"], connectionGeneration=group["connectionGeneration"],
                deviceTimestampUs=group["deviceTimestampUs"], hostReceivedUtc=group["hostTimestampUtc"],
                hostReceivedMonotonicNs=group["hostMonotonicNs"], frameKind=mode,
                rowsGeneration=group["rowsGeneration"], rowsRequestId=group["rowsRequestId"],
                modeGeneration=group["modeGeneration"], modeRequestId=group["modeRequestId"],
                profileGeneration=group["profileGeneration"], profileRequestId=group["profileRequestId"],
                configuredRowProfile=group["configuredRowProfile"], wireRowProfile=group["wireRowProfile"],
                rowModes=tuple(group["rowModes"]), rowUnits=tuple(group["rowUnits"]),
                rowScales=tuple(group["rowScales"]), rail=group["rail"],
            )
        )
    if not frames:
        raise ValueError("session file contains no frames")
    return frames


def _iter_frame_rows(frames: list[SessionFrame]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for frame_index, frame in enumerate(frames):
        values, raw = frame.physical_values(), frame.raw_fixed_values()
        valid, fresh = np.asarray(frame.valid, dtype=bool).reshape(64), frame.fresh_values()
        expected, acquired = frame.expected_values(), frame.acquired_values()
        expected_known, acquired_known = frame.expected_known_values(), frame.acquired_known_values()
        fresh_known = frame.fresh_known_values()
        error, errors = frame.error_values(), frame.error_code_values()
        error_reasons, pga, pga_bypass = frame.error_reason_values(), frame.pga_values(), frame.pga_bypass_values()
        row_modes, row_units, row_scales = frame.row_mode_values(), frame.row_unit_values(), frame.row_scale_values()
        rail = frame.rail or {}
        voltage_vss, voltage_normalised = _derived_voltage_values(frame)
        host_utc = frame.hostReceivedUtc
        if not host_utc and frame.hostWallTime is not None:
            from datetime import datetime, timezone

            host_utc = datetime.fromtimestamp(frame.hostWallTime, timezone.utc).isoformat()
        for row_index in range(8):
            for col_index in range(8):
                index = row_index * 8 + col_index
                physical = _blank_nan(values[index])
                raw_value = _blank_nan(raw[index])
                cell_mode = row_modes[row_index]
                unit = row_units[row_index]
                scale = row_scales[row_index]
                output.append(
                    {
                        "schemaVersion": SESSION_SCHEMA_VERSION,
                        "frameIndex": frame_index,
                        "sessionId": frame.sessionId,
                        "bootId": "" if frame.bootId is None else frame.bootId,
                        "connectionGeneration": frame.connectionGeneration,
                        "deviceTimestampUs": "" if frame.deviceTimestampUs is None else frame.deviceTimestampUs,
                        "hostTimestampUtc": host_utc,
                        "hostMonotonicNs": "" if frame.hostReceivedMonotonicNs is None else frame.hostReceivedMonotonicNs,
                        "seq": frame.seq,
                        "row": row_index + 1,
                        "col": col_index + 1,
                        "cell": _cell_label(row_index, col_index),
                        "mode": cell_mode,
                        "unit": unit,
                        "scale": "" if scale is None else scale,
                        "rawFixed": raw_value,
                        "physicalValue": physical,
                        "expected": bool(expected[index]),
                        "expectedKnown": bool(expected_known[index]),
                        "acquired": bool(acquired[index]),
                        "acquiredKnown": bool(acquired_known[index]),
                        "valid": bool(valid[index]),
                        "fresh": bool(fresh[index]),
                        "freshKnown": bool(fresh_known[index]),
                        "error": bool(error[index]),
                        "errorCode": "" if int(errors[index]) < 0 else int(errors[index]),
                        "errorReason": str(error_reasons[index]),
                        "pga": "" if int(pga[index]) < 0 else int(pga[index]),
                        "pgaBypass": bool(pga_bypass[index]),
                        "rowsGeneration": _blank_optional(frame.rowsGeneration),
                        "rowsRequestId": _blank_optional(frame.rowsRequestId),
                        "modeGeneration": _blank_optional(frame.modeGeneration),
                        "modeRequestId": _blank_optional(frame.modeRequestId),
                        "profileGeneration": _blank_optional(frame.profileGeneration),
                        "profileRequestId": _blank_optional(frame.profileRequestId),
                        "configuredRowProfile": _profile_text(frame.configuredRowProfile),
                        "wireRowProfile": frame.wireRowProfile or "",
                        "avdd": _rail_value(rail, "avddUv", "avdd"),
                        "avss": _rail_value(rail, "avssUv", "avss"),
                        "railSpan": _rail_value(rail, "railSpanUv", "railSpan"),
                        "railValid": bool(rail.get("railValid", False)),
                        "railFresh": bool(rail.get("railFresh", False)),
                        "railAge": _blank_optional(rail.get("railAge")),
                        "railSource": str(rail.get("railSource") or ""),
                        "railReason": str(rail.get("railReason") or ""),
                        "source": frame.source,
                        "measurementMode": frame.mode,
                        "timeSeconds": frame.timeSeconds,
                        "rows": frame.rows,
                        "generation": _blank_optional(frame.generation),
                        "requestId": _blank_optional(frame.requestId),
                        "correctedPf": physical if cell_mode == "CAP" else "",
                        "valueUv": raw_value if cell_mode == "VOLT" else "",
                        "valueV": physical if cell_mode == "VOLT" else "",
                        "valueMilliOhm": raw_value if cell_mode == "RES" else "",
                        "valueOhm": physical if cell_mode == "RES" else "",
                        "voltageVssRelativeV": _blank_nan(voltage_vss[index]) if cell_mode == "VOLT" else "",
                        "voltageRailNormalised": _blank_nan(voltage_normalised[index]) if cell_mode == "VOLT" else "",
                    }
                )
    return output


def _mode_value(value: Any) -> str:
    mode = str(value or "CAP").strip().upper()
    mode = {"CAPACITANCE": "CAP", "VOLTAGE": "VOLT", "RESISTANCE": "RES"}.get(mode, mode)
    if mode not in {"CAP", "VOLT", "RES", "MIXED"}:
        raise ValueError(f"unsupported measurement mode: {value}")
    return mode


def _mode_unit_scale(mode: str) -> tuple[str, int]:
    return {"CAP": ("pF", -6), "VOLT": ("V", -6), "RES": ("ohm", -3), "MIXED": ("", 0)}[
        _mode_value(mode)
    ]


def _bit_mask(values: np.ndarray, cells: int) -> int:
    mask = 0
    flat = np.asarray(values, dtype=bool).reshape(-1)
    if flat.size < cells:
        raise ValueError(f"bit mask needs {cells} values, got {flat.size}")
    for index, value in enumerate(flat[:cells]):
        if value:
            mask |= 1 << index
    return mask


def _device_row_mask(values: np.ndarray, rows: int, column_start: int, column_end: int) -> int:
    """Return a conservative CAP freshness mask for one four-channel FDC.

    Invalid values do not affect freshness.  Live CAP exports always contain
    one common bit for the four cells; requiring all four also avoids turning
    an unrepresentable hand-authored partial group into falsely fresh data.
    """

    mask = 0
    cells = np.asarray(values, dtype=bool).reshape(64)
    for row_index in range(rows):
        start = row_index * 8 + column_start
        end = row_index * 8 + column_end
        if bool(cells[start:end].all()):
            mask |= 1 << row_index
    return mask


def _flat_numeric_list(value: Any) -> np.ndarray:
    out = np.full(64, np.nan)
    if isinstance(value, (list, tuple, np.ndarray)):
        for index, item in enumerate(np.asarray(value, dtype=object).reshape(-1)[:64]):
            number = _json_number(item)
            if number is not None:
                out[index] = number
    return out


def _flat_bool_list(value: Any, default: np.ndarray) -> np.ndarray:
    if not isinstance(value, (list, tuple, np.ndarray)):
        return np.asarray(default, dtype=bool).reshape(64).copy()
    out = np.zeros(64, dtype=bool)
    flat = np.asarray(value, dtype=object).reshape(-1)
    for index, item in enumerate(flat[:64]):
        out[index] = bool(item)
    return out


def _flat_int_list(value: Any, fill: int) -> np.ndarray:
    out = np.full(64, fill, dtype=np.int16)
    if isinstance(value, (list, tuple, np.ndarray)):
        for index, item in enumerate(np.asarray(value, dtype=object).reshape(-1)[:64]):
            if item not in {None, ""}:
                out[index] = _int_value(item, fill)
    return out


def _flat_string_list(value: Any, fill: str) -> np.ndarray:
    out = np.full(64, fill, dtype=object)
    if isinstance(value, (list, tuple, np.ndarray)):
        for index, item in enumerate(np.asarray(value, dtype=object).reshape(-1)[:64]):
            if item is not None:
                out[index] = str(item)
    return out


def _matrix_from_json(value: Any, default: float = np.nan) -> np.ndarray:
    matrix = np.full((8, 8), default, dtype=np.float64)
    if not isinstance(value, list):
        return matrix
    for row_index, row in enumerate(value[:8]):
        if isinstance(row, list):
            for col_index, item in enumerate(row[:8]):
                number = _json_number(item)
                if number is not None:
                    matrix[row_index, col_index] = number
    return matrix


def _bool_matrix_from_json(value: Any) -> np.ndarray:
    matrix = np.zeros((8, 8), dtype=bool)
    if isinstance(value, list):
        for row_index, row in enumerate(value[:8]):
            if isinstance(row, list):
                for col_index, item in enumerate(row[:8]):
                    matrix[row_index, col_index] = bool(item)
    return matrix


def _flat_numeric_matrix(value: Any) -> np.ndarray:
    return _matrix_from_json(value).reshape(64)


def _flat_bool_matrix(value: Any, default: np.ndarray) -> np.ndarray:
    return _bool_matrix_from_json(value).reshape(64) if isinstance(value, list) else np.asarray(default, dtype=bool).reshape(64)


def _flat_int_matrix(value: Any, fill: int) -> np.ndarray:
    if not isinstance(value, list):
        return np.full(64, fill, dtype=np.int16)
    out = np.full((8, 8), fill, dtype=np.int16)
    for row_index, row in enumerate(value[:8]):
        if isinstance(row, list):
            for col_index, item in enumerate(row[:8]):
                if item is not None:
                    out[row_index, col_index] = _int_value(item, fill)
    return out.reshape(64)


def _flat_string_matrix(value: Any, fill: str) -> np.ndarray:
    if not isinstance(value, list):
        return np.full(64, fill, dtype=object)
    out = np.full((8, 8), fill, dtype=object)
    for row_index, row in enumerate(value[:8]):
        if isinstance(row, list):
            for col_index, item in enumerate(row[:8]):
                if item is not None:
                    out[row_index, col_index] = str(item)
    return out.reshape(64)


def _h5_optional(handle: Any, name: str) -> Any:
    return handle[name][()] if name in handle else None


def _h5_group_dict(handle: Any, name: str) -> dict[str, Any]:
    if name not in handle:
        return {}
    return {str(key): item[()] for key, item in handle[name].items()}


def _array_matrix(value: Any, index: int, default: Any) -> np.ndarray:
    fallback = np.asarray(default if np.asarray(default).shape == (8, 8) else np.full((8, 8), default)).reshape(64)
    if value is None:
        return fallback
    array = np.asarray(value)
    # scipy.io.loadmat represents an all-empty MATLAB char array as (0, 0).
    # Empty error-reason matrices are semantically the supplied default, not
    # a missing frame, and must not make an otherwise lossless session fail.
    if array.size == 0 or array.ndim == 0 or index >= array.shape[0]:
        return fallback
    selected = array[index]
    if selected.size != 64:
        return fallback
    return selected.reshape(64)


def _array_mode(value: Any, index: int) -> str:
    if value is None:
        return "CAP"
    flat = np.asarray(value).reshape(-1)
    raw = flat[index] if index < flat.size else "CAP"
    if isinstance(raw, bytes):
        raw = raw.decode("ascii", errors="ignore")
    raw_text = str(raw).strip("[]' ")
    return _mode_value(raw_text or "CAP")


def _array_string(value: Any, index: int, default: str) -> str:
    if value is None:
        return default
    flat = np.asarray(value).reshape(-1)
    raw = flat[index] if index < flat.size else default
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8", errors="replace")
    text = str(raw).strip("[]' ")
    return text or default


def _array_int(value: Any, index: int, default: int) -> int:
    if value is None:
        return default
    flat = np.asarray(value).reshape(-1)
    return _int_value(flat[index] if index < flat.size else default, default)


def _array_optional_int(value: Any, index: int) -> int | None:
    parsed = _array_int(value, index, -1)
    return None if parsed < 0 else parsed


def _array_optional_string(value: Any, index: int) -> str | None:
    text = _array_string(value, index, "")
    return text or None


def _array_profile(value: Any, index: int, *, allow_none: bool) -> tuple[str, ...] | None:
    text = _array_optional_string(value, index)
    return _optional_profile(text, allow_none=allow_none) if text else None


def _array_row_strings(value: Any, index: int) -> tuple[str, ...] | None:
    if value is None:
        return None
    array = np.asarray(value)
    if array.ndim < 2 or index >= array.shape[0]:
        return None
    output: list[str] = []
    for raw in array[index].reshape(-1)[:8]:
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")
        output.append(str(raw).strip("[]' "))
    return tuple(output) if len(output) == 8 else None


def _array_row_scales(value: Any, index: int) -> tuple[int | None, ...] | None:
    if value is None:
        return None
    array = np.asarray(value)
    if array.ndim < 2 or index >= array.shape[0]:
        return None
    values = tuple(int(raw) for raw in array[index].reshape(-1)[:8])
    if len(values) != 8:
        return None
    return tuple(None if item == -128 else item for item in values)


def _extra_value(extras: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in extras:
            return extras[name]
    return None


def _json_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if np.isfinite(number) else None


def _blank_nan(value: Any) -> float | str:
    number = _json_number(value)
    return "" if number is None else number


def _bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _optional_int(value: Any) -> int | None:
    if value in {None, ""}:
        return None
    return _int_value(value, 0)


def _optional_float(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    return _float_value(value, 0.0)


def _optional_string(value: Any) -> str | None:
    if value in {None, ""}:
        return None
    return str(value)


def _int_value(value: Any, default: int) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _float_value(value: Any, default: float) -> float:
    number = _json_number(value)
    return default if number is None else number


def _rows_value(value: Any, default: int) -> int:
    rows = _int_value(value, default)
    if not 1 <= rows <= 8:
        raise ValueError("rows must be 1..8")
    return rows


def _row_index(value: Any) -> int:
    index = _int_value(value, 0) - 1
    if not 0 <= index < 8:
        raise ValueError("row must be 1..8")
    return index


def _col_index(value: Any) -> int:
    index = _int_value(value, 0) - 1
    if not 0 <= index < 8:
        raise ValueError("col must be 1..8")
    return index


def _cell_label(row_index: int, col_index: int) -> str:
    return f"S{row_index + 1}D{col_index + 1}"


def _string_value(value: Any) -> str:
    return "" if value is None else str(value)


def _blank_optional(value: Any) -> Any:
    return "" if value is None else value


def _rail_value(rail: dict[str, Any], primary: str, alias: str) -> Any:
    value = rail.get(primary)
    if value is None:
        value = rail.get(alias)
    return "" if value is None else value


def _profile_text(profile: tuple[str, ...] | None) -> str:
    if profile is None:
        return ""
    return _configured_profile_wire(profile)


def _array_int_or_missing(value: Any) -> int:
    return -1 if value is None or value == "" else int(value)


def _array_signed_int_or_missing(value: Any) -> int:
    return SIGNED_INT_MISSING if value is None or value == "" else int(value)


def _array_optional_signed_int(value: Any, index: int) -> int | None:
    if value is None:
        return None
    flat = np.asarray(value).reshape(-1)
    if index >= flat.size:
        return None
    parsed = int(flat[index])
    return None if parsed == SIGNED_INT_MISSING else parsed


def _derived_voltage_values(frame: SessionFrame) -> tuple[np.ndarray, np.ndarray]:
    relative = np.full(64, np.nan, dtype=np.float64)
    normalised = np.full(64, np.nan, dtype=np.float64)
    rail = frame.rail or {}
    avdd = _optional_int(rail.get("avddUv") if rail.get("avddUv") is not None else rail.get("avdd"))
    avss = _optional_int(rail.get("avssUv") if rail.get("avssUv") is not None else rail.get("avss"))
    rail_boot = _optional_int(rail.get("bootId"))
    same_boot = frame.bootId is not None and rail_boot == frame.bootId
    usable_rail = bool(
        same_boot
        and rail.get("railValid")
        and rail.get("railFresh")
        and avdd is not None
        and avss is not None
        and avdd > avss
    )
    if not usable_rail:
        return relative, normalised
    values = frame.physical_values()
    valid = np.asarray(frame.valid, dtype=bool).reshape(64)
    fresh = frame.fresh_values()
    row_modes = frame.row_mode_values()
    for index, value in enumerate(values):
        if row_modes[index // 8] != "VOLT" or not valid[index] or not fresh[index] or not np.isfinite(value):
            continue
        relative[index] = float(value) - float(avss) * 1e-6
        normalised[index] = relative[index] / ((float(avdd) - float(avss)) * 1e-6)
    return relative, normalised


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _frame_metadata_row(frame_index: int, frame: SessionFrame) -> dict[str, Any]:
    rail = frame.rail or {}
    return {
        "frameIndex": frame_index,
        "sessionId": frame.sessionId,
        "bootId": _blank_optional(frame.bootId),
        "connectionGeneration": frame.connectionGeneration,
        "deviceTimestampUs": _blank_optional(frame.deviceTimestampUs),
        "hostTimestampUtc": frame.hostReceivedUtc,
        "hostMonotonicNs": _blank_optional(frame.hostReceivedMonotonicNs),
        "seq": frame.seq,
        "rows": frame.rows,
        "frameKind": frame.mode,
        "rowsGeneration": _blank_optional(frame.rowsGeneration),
        "rowsRequestId": _blank_optional(frame.rowsRequestId),
        "modeGeneration": _blank_optional(frame.modeGeneration),
        "modeRequestId": _blank_optional(frame.modeRequestId),
        "profileGeneration": _blank_optional(frame.profileGeneration),
        "profileRequestId": _blank_optional(frame.profileRequestId),
        "configuredRowProfile": _profile_text(frame.configuredRowProfile),
        "wireRowProfile": frame.wireRowProfile or "",
        "railValid": bool(rail.get("railValid", False)),
        "railFresh": bool(rail.get("railFresh", False)),
        "railAge": _blank_optional(rail.get("railAge")),
        "avdd": _rail_value(rail, "avddUv", "avdd"),
        "avss": _rail_value(rail, "avssUv", "avss"),
        "railSpan": _rail_value(rail, "railSpanUv", "railSpan"),
        "railSource": str(rail.get("railSource") or ""),
        "railReason": str(rail.get("railReason") or ""),
        "source": frame.source,
    }


def _append_wide_sheet(
    sheet: Any,
    frames: list[SessionFrame],
    mode: str,
    *,
    value_kind: str = "ground",
) -> None:
    identity = [
        "schemaVersion", "frameIndex", "sessionId", "bootId", "connectionGeneration",
        "deviceTimestampUs", "hostTimestampUtc", "hostMonotonicNs", "seq", "rows", "frameKind",
    ]
    cells = [_cell_label(row, col) for row in range(8) for col in range(8)]
    sheet.append(identity + cells)
    for frame_index, frame in enumerate(frames):
        values = _wide_values(frame, value_kind)
        valid = np.asarray(frame.valid, dtype=bool).reshape(64)
        row_modes = frame.row_mode_values()
        row: list[Any] = [
            SESSION_SCHEMA_VERSION,
            frame_index,
            frame.sessionId,
            _blank_optional(frame.bootId),
            frame.connectionGeneration,
            _blank_optional(frame.deviceTimestampUs),
            frame.hostReceivedUtc,
            _blank_optional(frame.hostReceivedMonotonicNs),
            frame.seq,
            frame.rows,
            frame.mode,
        ]
        for index, value in enumerate(values):
            row_index = index // 8
            usable = row_index < frame.rows and row_modes[row_index] == mode and bool(valid[index])
            row.append(_json_number(value) if usable else None)
        sheet.append(row)


def _wide_values(frame: SessionFrame, value_kind: str) -> np.ndarray:
    if value_kind == "ground":
        return frame.physical_values()
    relative, normalised = _derived_voltage_values(frame)
    if value_kind == "vss_relative":
        return relative
    if value_kind == "rail_normalized":
        return normalised
    raise ValueError(f"unsupported wide value kind: {value_kind}")


def _wide_csv(frames: list[SessionFrame], mode: str, *, value_kind: str = "ground") -> bytes:
    output = io.StringIO(newline="")
    identity = [
        "schemaVersion", "frameIndex", "sessionId", "bootId", "connectionGeneration",
        "deviceTimestampUs", "hostTimestampUtc", "hostMonotonicNs", "seq", "rows", "frameKind",
    ]
    cells = [_cell_label(row, col) for row in range(8) for col in range(8)]
    writer = csv.writer(output)
    writer.writerow(identity + cells)
    for frame_index, frame in enumerate(frames):
        values = _wide_values(frame, value_kind)
        valid = np.asarray(frame.valid, dtype=bool).reshape(64)
        fresh = frame.fresh_values()
        row_modes = frame.row_mode_values()
        row: list[Any] = [
            SESSION_SCHEMA_VERSION,
            frame_index,
            frame.sessionId,
            _blank_optional(frame.bootId),
            frame.connectionGeneration,
            _blank_optional(frame.deviceTimestampUs),
            frame.hostReceivedUtc,
            _blank_optional(frame.hostReceivedMonotonicNs),
            frame.seq,
            frame.rows,
            frame.mode,
        ]
        for index, value in enumerate(values):
            row_index = index // 8
            usable = (
                row_index < frame.rows
                and row_modes[row_index] == mode
                and bool(valid[index])
                and bool(fresh[index])
                and np.isfinite(value)
            )
            row.append(float(value) if usable else "")
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _key_value_csv(values: dict[str, Any]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["key", "value"])
    for key, value in values.items():
        writer.writerow([key, _string_value(value)])
    return output.getvalue().encode("utf-8")


def _frame_metadata_csv(frames: list[SessionFrame]) -> bytes:
    columns = [
        "frameIndex", "sessionId", "bootId", "connectionGeneration", "deviceTimestampUs",
        "hostTimestampUtc", "hostMonotonicNs", "seq", "rows", "frameKind", "rowsGeneration",
        "rowsRequestId", "modeGeneration", "modeRequestId", "profileGeneration", "profileRequestId",
        "configuredRowProfile", "wireRowProfile", "railValid", "railFresh", "railAge", "avdd",
        "avss", "railSpan", "railSource", "railReason", "source",
    ]
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for frame_index, frame in enumerate(frames):
        writer.writerow(_frame_metadata_row(frame_index, frame))
    return output.getvalue().encode("utf-8")


def _mapping_csv(values: list[dict[str, Any]], fallback: tuple[str, ...]) -> bytes:
    keys = sorted({str(key) for value in values for key in value}) or list(fallback)
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=keys, extrasaction="ignore")
    writer.writeheader()
    for value in values:
        writer.writerow(
            {
                key: _json_text(value.get(key)) if isinstance(value.get(key), (dict, list, tuple)) else value.get(key)
                for key in keys
            }
        )
    return output.getvalue().encode("utf-8")


def _offsets_csv(offsets: np.ndarray) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["row", "col", "offsetPf"])
    for row in range(8):
        for col in range(8):
            writer.writerow([row + 1, col + 1, float(offsets[row, col])])
    return output.getvalue().encode("utf-8")


def _append_mapping_sheet(sheet: Any, values: list[dict[str, Any]]) -> None:
    keys = sorted({str(key) for value in values for key in value})
    if not keys:
        sheet.append(["kind", "hostTimestampUtc"])
        return
    sheet.append(keys)
    for value in values:
        sheet.append(
            [
                _json_text(value.get(key)) if isinstance(value.get(key), (dict, list, tuple)) else value.get(key)
                for key in keys
            ]
        )


def _known_values(value: np.ndarray | None, default: bool) -> np.ndarray:
    if value is None:
        return np.full(64, bool(default), dtype=bool)
    return np.asarray(value, dtype=bool).reshape(64)


def _profile_tuple(value: Any, *, allow_none: bool) -> tuple[str, ...]:
    if isinstance(value, str):
        aliases = {"C": "CAP", "V": "VOLT", "R": "RES", "N": "NONE"}
        values = tuple(aliases.get(char, char) for char in value.strip().upper())
    else:
        try:
            values = tuple(str(item).strip().upper() for item in value)
        except TypeError as exc:
            raise ValueError("row profile must contain exactly 8 values") from exc
    if len(values) != 8:
        raise ValueError("row profile must contain exactly 8 values")
    allowed = {"CAP", "VOLT", "RES", "NONE"} if allow_none else {"CAP", "VOLT", "RES"}
    if any(item not in allowed for item in values):
        raise ValueError("row profile contains an unsupported mode")
    return values


def _optional_profile(value: Any, *, allow_none: bool = True) -> tuple[str, ...] | None:
    if value is None or (isinstance(value, str) and value == ""):
        return None
    return _profile_tuple(value, allow_none=allow_none)


def _string_tuple(value: Any, default: str) -> tuple[str, ...]:
    try:
        values = tuple(str(item) if item is not None else default for item in value)
    except TypeError as exc:
        raise ValueError("row metadata must contain exactly 8 values") from exc
    if len(values) != 8:
        raise ValueError("row metadata must contain exactly 8 values")
    return values


def _optional_string_tuple(value: Any) -> tuple[str, ...] | None:
    if value is None:
        return None
    return _string_tuple(value, "")


def _optional_scale_tuple(value: Any) -> tuple[int | None, ...] | None:
    if value is None:
        return None
    try:
        values = tuple(None if item is None or item == "" else int(item) for item in value)
    except (TypeError, ValueError) as exc:
        raise ValueError("row scale metadata must contain integers or null") from exc
    if len(values) != 8:
        raise ValueError("row scale metadata must contain exactly 8 values")
    return values


def _wire_acquisition_masks(frame: SessionFrame, cells: int) -> tuple[np.ndarray, np.ndarray]:
    expected = frame.expected_values()
    acquired = frame.acquired_values()
    expected_known = frame.expected_known_values()
    acquired_known = frame.acquired_known_values()
    wire_expected = expected & expected_known
    wire_acquired = acquired & acquired_known
    # Acquired implies expected in the 8045 contract.  If a legacy migration
    # knows acquisition but not the expected bit, this implication is factual
    # and is the only safe representable wire form.
    wire_expected |= wire_acquired
    active = np.zeros(64, dtype=bool)
    active[:cells] = True
    return wire_expected & active, wire_acquired & active


def _configured_profile_wire(profile: tuple[str, ...]) -> str:
    return "".join({"CAP": "C", "VOLT": "V", "RES": "R"}[mode] for mode in _profile_tuple(profile, allow_none=False))


def _mixed_wire_profile(row_modes: tuple[str, ...], rows: int) -> str:
    normalized = _profile_tuple(row_modes, allow_none=True)
    chars = {"CAP": "C", "VOLT": "V", "RES": "R", "NONE": "N"}
    profile = tuple(normalized[index] if index < rows else "NONE" for index in range(8))
    if any(profile[index] == "NONE" for index in range(rows)):
        raise ValueError("active Mixed rows cannot use NONE")
    return "".join(chars[mode] for mode in profile)
